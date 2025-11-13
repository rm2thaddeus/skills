import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional, Tuple

import pandas as pd


def find_latest_op(artifacts_root: Path) -> Optional[Path]:
    ops = [p for p in artifacts_root.glob("*/") if p.is_dir()]
    if not ops:
        return None
    return sorted(ops, key=lambda p: p.stat().st_mtime, reverse=True)[0]


def load_minimal_datasets(op_dir: Path) -> Tuple[pd.DataFrame, pd.DataFrame]:
    enriched = op_dir / "output" / "enriched"
    i_path = enriched / "interviews_dataset.xlsx"
    a_path = enriched / "applications_dataset.xlsx"
    interviews = pd.read_excel(i_path) if i_path.exists() else pd.DataFrame()
    applications = pd.read_excel(a_path) if a_path.exists() else pd.DataFrame()
    if not interviews.empty:
        interviews["event_datetime"] = pd.to_datetime(interviews.get("event_datetime"), errors="coerce")
        if "event_year" not in interviews:
            interviews["event_year"] = interviews["event_datetime"].dt.year
    if not applications.empty:
        applications["applied_date"] = pd.to_datetime(applications.get("applied_date"), errors="coerce")
        if "year" not in applications:
            applications["year"] = applications["applied_date"].dt.year
    return interviews, applications


def load_manual_periods(config_path: Optional[str]) -> Dict[str, list]:
    if not config_path:
        return {}
    p = Path(config_path)
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def inter_application_intervals_days(apps: pd.DataFrame, year: int) -> Tuple[Optional[float], Optional[float]]:
    sub = apps[apps["year"] == year].copy()
    if sub.empty:
        return None, None
    dates = sorted(sub["applied_date"].dropna().dt.floor("D").tolist())
    if len(dates) < 2:
        return None, None
    diffs = pd.Series(pd.Series(dates).diff().dropna().dt.days)
    return float(diffs.median()), float(diffs.mean())


def median_days_to_next_event(apps: pd.DataFrame, events: pd.Series, year: int) -> Optional[float]:
    a = apps[apps["year"] == year][["applied_date"]].dropna().copy()
    if a.empty:
        return None
    a = a.sort_values("applied_date")
    e = pd.Series(events.dropna()).sort_values()
    if e.empty:
        return None
    # For each application date, find next event date (forward asof)
    a["__key__"] = a["applied_date"]
    e_df = pd.DataFrame({"event_date": e})
    m = pd.merge_asof(a.sort_values("applied_date"), e_df.sort_values("event_date"), left_on="applied_date", right_on="event_date", direction="forward")
    if m["event_date"].isna().all():
        return None
    deltas = (m["event_date"] - m["applied_date"]).dt.days.dropna()
    if deltas.empty:
        return None
    return float(deltas.median())


def is_role_related(text: Optional[str]) -> bool:
    if text is None:
        return False
    t = str(text).lower()
    keywords = [
        "role", "position", "opening", "opportunity", "hiring", "recruit", "talent",
        "product", "engineer", "developer", "data", "ml", "manager", "designer"
    ]
    return any(kw in t for kw in keywords)


def detect_application_periods(apps: pd.DataFrame, year: int, gap_days: int = 7) -> pd.DataFrame:
    sub = apps[apps["year"] == year][["applied_date"]].dropna().copy()
    if sub.empty:
        return pd.DataFrame(columns=["period_index","start_date","end_date","num_days","num_applications","applications_per_day"])
    dates = sorted(sub["applied_date"].dt.floor("D").tolist())
    periods = []
    cur = [dates[0]]
    for d in dates[1:]:
        if (d - cur[-1]).days > gap_days:
            periods.append(cur)
            cur = [d]
        else:
            cur.append(d)
    periods.append(cur)
    rows = []
    for i, grp in enumerate(periods, start=1):
        start = grp[0]
        end = grp[-1]
        num_days = max(1, (end - start).days + 1)
        num_apps = len(grp)
        rows.append({
            "period_index": i,
            "start_date": start.date().isoformat(),
            "end_date": end.date().isoformat(),
            "num_days": num_days,
            "num_applications": num_apps,
            "applications_per_day": round(num_apps / num_days, 4),
        })
    return pd.DataFrame(rows)


def compute_period_metrics(apps: pd.DataFrame, interviews: pd.DataFrame, period: dict) -> dict:
    start = pd.to_datetime(period.get("start_date"), errors="coerce")
    end = pd.to_datetime(period.get("end_date"), errors="coerce")
    name = period.get("name")
    if pd.isna(start) or pd.isna(end):
        return {"name": name, "error": "invalid_dates"}
    # Filter windows
    a = apps[(apps["applied_date"] >= start) & (apps["applied_date"] <= end)].copy()
    iv = interviews[(interviews["event_datetime"] >= start) & (interviews["event_datetime"] <= end)].copy()
    # Counts
    total_apps = int(len(a))
    total_interviews = int(((iv.get("event_type").fillna("") == "interview")).sum()) if not iv.empty else 0
    total_outreach = int(((iv.get("event_type").fillna("") == "outreach")).sum()) if not iv.empty else 0
    # Inbound role contacts
    inbound = iv[(iv.get("event_type").fillna("") == "outreach")]
    if "direction" in inbound.columns:
        inbound = inbound[inbound["direction"].fillna("").str.lower() == "inbound"]
    def _role_related(row) -> bool:
        r = str(row.get("role") or "").strip()
        if r:
            return True
        n = str(row.get("notes") or "").lower()
        for kw in ("role","position","opening","opportunity","hiring","recruit","product","engineer","developer","data","ml","manager","designer"):
            if kw in n:
                return True
        return False
    inbound_role = inbound[inbound.apply(_role_related, axis=1)] if not inbound.empty else inbound
    inbound_role_contacts = int(len(inbound_role)) if inbound_role is not None else 0
    # Intensity
    num_days = max(1, int((end - start).days) + 1)
    apps_per_day = (total_apps / num_days) if num_days else None
    conversion = (total_interviews / total_apps) if total_apps else None
    return {
        "name": name,
        "start_date": start.date().isoformat(),
        "end_date": end.date().isoformat(),
        "num_days": num_days,
        "total_applications": total_apps,
        "applications_per_day": apps_per_day,
        "total_interviews": total_interviews,
        "total_outreach": total_outreach,
        "inbound_role_contacts": inbound_role_contacts,
        "interviews_per_application": conversion,
        "notes": period.get("notes"),
    }


def compute_metrics(op_dir: Path, periods_config: Optional[str] = None) -> Dict[str, Dict[str, object]]:
    interviews, applications = load_minimal_datasets(op_dir)
    manual_periods_cfg = load_manual_periods(periods_config)
    metrics: Dict[str, Dict[str, object]] = {}
    years = [y for y in (2022, 2025)]
    for y in years:
        total_apps = int((applications.get("year") == y).sum()) if not applications.empty else 0
        evt_year = interviews.get("event_year") if not interviews.empty else None
        evt_type = interviews.get("event_type") if not interviews.empty else None
        direction = interviews.get("direction") if not interviews.empty else None
        evt_year_mask = (evt_year == y) if evt_year is not None else pd.Series([], dtype=bool)
        type_interview_mask = (evt_type.fillna("") == "interview") if evt_type is not None else pd.Series([], dtype=bool)
        type_outreach_mask = (evt_type.fillna("") == "outreach") if evt_type is not None else pd.Series([], dtype=bool)
        direction_inbound_mask = (direction.fillna("").str.lower() == "inbound") if direction is not None else pd.Series([], dtype=bool)

        total_interviews = int((evt_year_mask & type_interview_mask).sum()) if not interviews.empty else 0
        total_outreach = int((evt_year_mask & type_outreach_mask).sum()) if not interviews.empty else 0
        inbound_outreach = int((evt_year_mask & type_outreach_mask & direction_inbound_mask).sum()) if not interviews.empty else 0

        med_gap, mean_gap = inter_application_intervals_days(applications, y)

        # Median days from application to next interview and to next inbound outreach
        interview_dates = interviews[evt_year_mask & type_interview_mask]["event_datetime"] if not interviews.empty else pd.Series([], dtype="datetime64[ns]")
        inbound_outreach_df = interviews[evt_year_mask & type_outreach_mask & direction_inbound_mask].copy() if not interviews.empty else pd.DataFrame()
        inbound_dates = inbound_outreach_df["event_datetime"] if not inbound_outreach_df.empty else pd.Series([], dtype="datetime64[ns]")

        # Role-related inbound contacts metric (uses role column or notes text)
        inbound_role_mask = pd.Series([], dtype=bool)
        if not inbound_outreach_df.empty:
            role_col = inbound_outreach_df.get("role")
            notes_col = inbound_outreach_df.get("notes")
            role_has = role_col.notna() & role_col.astype(str).str.strip().ne("") if role_col is not None else pd.Series([False]*len(inbound_outreach_df))
            notes_has = notes_col.apply(is_role_related) if notes_col is not None else pd.Series([False]*len(inbound_outreach_df))
            inbound_role_mask = role_has | notes_has
        inbound_role_contacts = int(inbound_role_mask.sum()) if not inbound_outreach_df.empty else 0
        inbound_role_rate_per_app = (inbound_role_contacts / total_apps) if total_apps else None

        inbound_role_dates = inbound_outreach_df.loc[inbound_role_mask, "event_datetime"] if not inbound_outreach_df.empty else pd.Series([], dtype="datetime64[ns]")
        med_days_to_inbound_role = median_days_to_next_event(applications, inbound_role_dates, y)
        med_days_to_interview = median_days_to_next_event(applications, interview_dates, y)
        med_days_to_inbound = median_days_to_next_event(applications, inbound_dates, y)

        # Active days with applications
        daily_apps = applications[applications["year"] == y]["applied_date"].dropna().dt.floor("D").value_counts()
        active_days = int((daily_apps > 0).sum()) if not daily_apps.empty else 0
        intensity_apps_per_active_day = (total_apps / active_days) if active_days else None

        conversion = (total_interviews / total_apps) if total_apps else None

        # Application periods segmentation (7-day gap)
        periods_df = detect_application_periods(applications, y, gap_days=7)
        periods = periods_df.to_dict(orient="records") if not periods_df.empty else []
        top_period = None
        if periods:
            top_period = max(periods, key=lambda r: r.get("applications_per_day") or 0)

        entry = {
            "total_applications": total_apps,
            "total_interviews": total_interviews,
            "total_outreach": total_outreach,
            "inbound_outreach": inbound_outreach,
            "inbound_role_contacts": inbound_role_contacts,
            "inbound_role_contacts_per_application": inbound_role_rate_per_app,
            "applications_per_active_day": intensity_apps_per_active_day,
            "median_days_between_applications": med_gap,
            "mean_days_between_applications": mean_gap,
            "median_days_to_next_interview": med_days_to_interview,
            "median_days_to_next_inbound": med_days_to_inbound,
            "median_days_to_next_inbound_role": med_days_to_inbound_role,
            "interviews_per_application": conversion,
            "application_periods_count": len(periods),
            "application_periods": periods,
            "top_period_by_intensity": top_period,
        }
        # Manual vs AI-assisted assumptions: 2025 apps are AI-assisted; 2022 apps are manual
        if y == 2025:
            entry["ai_assisted_applications"] = total_apps
            entry["manual_applications"] = 0
            entry["ai_assisted_ratio"] = 1.0
        elif y == 2022:
            entry["ai_assisted_applications"] = 0
            entry["manual_applications"] = total_apps
            entry["ai_assisted_ratio"] = 0.0
        # Manual periods override/additional segmentation
        year_periods = manual_periods_cfg.get(str(y)) or manual_periods_cfg.get(y)
        if year_periods:
            per_metrics = []
            for pdef in year_periods:
                per_metrics.append(compute_period_metrics(applications, interviews, pdef))
            entry["manual_periods"] = year_periods
            entry["manual_periods_metrics"] = per_metrics
        metrics[str(y)] = entry

    return metrics


def write_metrics(op_dir: Path, metrics: Dict[str, Dict[str, object]]) -> Path:
    report_dir = op_dir / "output" / "report"
    report_dir.mkdir(parents=True, exist_ok=True)
    # Write JSON
    (report_dir / "metrics_summary.json").write_text(json.dumps(metrics, indent=2), encoding="utf-8")
    # Write Excel with PROVENANCE last
    xlsx = report_dir / "metrics_summary.xlsx"
    rows = []
    years = sorted(metrics.keys())
    all_keys = sorted(set(k for y in years for k in metrics[y].keys()))
    for k in all_keys:
        row = {"metric": k}
        for y in years:
            row[y] = metrics[y].get(k)
        # delta 2025-2022 when applicable numeric
        try:
            v22 = metrics.get("2022", {}).get(k)
            v25 = metrics.get("2025", {}).get(k)
            if isinstance(v22, (int, float)) and isinstance(v25, (int, float)):
                row["delta_2025_minus_2022"] = v25 - v22
        except Exception:
            pass
        rows.append(row)
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(xlsx, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Metrics")
    # Append provenance last
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(xlsx))
        if "PROVENANCE" in wb.sheetnames:
            wb.remove(wb["PROVENANCE"])
        ws = wb.create_sheet("PROVENANCE")
        ws["A1"] = "operation_id"; ws["B1"] = op_dir.name
        ws["A2"] = "generated_at_utc"; ws["B2"] = datetime.utcnow().isoformat() + "Z"
        ws["A3"] = "output_path"; ws["B3"] = str(xlsx)
        wb.save(str(xlsx))
    except Exception:
        pass
    return xlsx


def main() -> None:
    ap = argparse.ArgumentParser(description="Derive comparison metrics for 2022 vs 2025")
    ap.add_argument("--artifacts-root", default="artifacts/operations", help="Artifacts root")
    ap.add_argument("--op-dir", default=None, help="Specific operation dir")
    ap.add_argument("--periods-config", default="config/periods.json", help="Optional JSON config for manual periods")
    args = ap.parse_args()

    artifacts_root = Path(args.artifacts_root)
    op_dir = Path(args.op_dir) if args.op_dir else find_latest_op(artifacts_root)
    if not op_dir or not op_dir.exists():
        print("No operation directory found.")
        return

    metrics = compute_metrics(op_dir, periods_config=args.periods_config)
    xlsx = write_metrics(op_dir, metrics)
    out = {
        "operation_id": op_dir.name,
        "metrics": metrics,
        "metrics_xlsx": str(xlsx),
    }
    print(json.dumps(out, indent=2))


if __name__ == "__main__":
    main()


