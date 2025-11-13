import argparse
import csv
import json
import os
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


# -----------------------------
# Utilities
# -----------------------------

def ensure_op_dir(output_root: Path) -> Path:
    op_id = datetime.utcnow().strftime("%Y%m%d-%H%M%S-%f")
    op = output_root / op_id
    (op / "input").mkdir(parents=True, exist_ok=True)
    (op / "processing").mkdir(parents=True, exist_ok=True)
    (op / "output").mkdir(parents=True, exist_ok=True)
    (op / "output" / "enriched").mkdir(parents=True, exist_ok=True)
    (op / "output" / "report").mkdir(parents=True, exist_ok=True)
    (op / "output" / "charts").mkdir(parents=True, exist_ok=True)
    return op


def write_audit(op: Path, agent: str, inputs: Dict[str, str], note: str) -> None:
    audit = {
        "operation_id": op.name,
        "timestamp": datetime.utcnow().isoformat() + "Z",
        "agent": agent,
        "inputs": inputs,
        "output_dir": str(op / "output"),
        "changes": [],
        "validation": {"status": "success", "notes": note},
        "approval_status": "pending",
    }
    (op / "audit.json").write_text(json.dumps(audit, indent=2), encoding="utf-8")


def robust_read_csv(path: Path) -> pd.DataFrame:
    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
                sep = dialect.delimiter
            except Exception:
                sep = ","
        return pd.read_csv(path, encoding="utf-8-sig", sep=sep, dtype=str)
    except Exception:
        try:
            return pd.read_csv(path, encoding="utf-8", engine="python", dtype=str, on_bad_lines="skip")
        except Exception:
            return pd.read_csv(path, encoding="latin1", engine="python", dtype=str, on_bad_lines="skip")


def parse_time_range(text: Optional[str]) -> Tuple[Optional[str], Optional[str], Optional[int]]:
    if not isinstance(text, str):
        return None, None, None
    m = re.search(r"(\d{1,2}:\d{2})\D+(\d{1,2}:\d{2})", text)
    start = m.group(1) if m else None
    end = m.group(2) if m else None
    dur = None
    m2 = re.search(r"(\d+)\s*min", text, flags=re.I)
    if m2:
        try:
            dur = int(m2.group(1))
        except Exception:
            dur = None
    return start, end, dur


def extract_dates_from_text(text: Optional[str]) -> List[str]:
    if not isinstance(text, str) or not text.strip():
        return []
    s = text.replace("\u2013", "-").replace("\u2014", "-")
    # find day-month-year mentions like "Apr 17, 2025" or "September 30, 2025"
    month_names = (
        r"Jan|January|Feb|February|Mar|March|Apr|April|May|Jun|June|Jul|July|Aug|August|Sep|Sept|September|Oct|October|Nov|November|Dec|December"
    )
    import re as _re
    out: List[str] = []
    for m in _re.finditer(fr"\b({month_names})\s+\d{{1,2}},\s*\d{{4}}", s, flags=_re.I):
        try:
            dt = pd.to_datetime(m.group(0), errors="coerce")
            if pd.notna(dt):
                out.append(pd.Timestamp(dt).date().isoformat())
        except Exception:
            continue
    # If none found, try month-year (assume day 1)
    if not out:
        for m in _re.finditer(fr"\b({month_names})\s+\d{{4}}\b", s, flags=_re.I):
            try:
                dt = pd.to_datetime("1 " + m.group(0), errors="coerce")
                if pd.notna(dt):
                    out.append(pd.Timestamp(dt).date().isoformat())
            except Exception:
                continue
    return list(dict.fromkeys(out))

def infer_role_from_text(text: Optional[str]) -> Optional[str]:
    if not isinstance(text, str) or not text.strip():
        return None
    t = text.lower()
    patterns = [
        (r"product\s+manager|pm\b", "Product Manager"),
        (r"data\s+scientist|ds\b", "Data Scientist"),
        (r"data\s+engineer", "Data Engineer"),
        (r"machine\s+learning\s+engineer|ml\s+engineer", "ML Engineer"),
        (r"software\s+engineer|developer|frontend|backend|full[- ]stack", "Software Engineer"),
        (r"design(er)?|ux|ui", "Designer"),
        (r"recruiter|talent\s+acquisition|ta\b", "Recruiter"),
        (r"hr\b|human\s+resources", "HR"),
        (r"cto\b|chief\s+technology", "CTO"),
        (r"ceo\b|chief\s+executive", "CEO"),
        (r"founder|co[- ]founder", "Founder"),
        (r"product\s+owner", "Product Owner"),
        (r"analyst", "Analyst"),
        (r"manager", "Manager"),
    ]
    for pat, label in patterns:
        if re.search(pat, t, flags=re.I):
            return label
    return None


# -----------------------------
# Schema map (stable)
# -----------------------------

SCHEMA = {
    "interview_dataset_2022.csv": {
        "canonical": {
            "date": "Date",
            "start_time": "Start Time",
            "end_time": "End Time",
            "company": "Company",
            "role": "Role",
            "medium": "Medium",
            "source": "Source",
            "contacts": "Contact",
            "notes": "Notes",
        }
    },
    "interview_dataset.csv": {
        "canonical": {
            "date": "Date (2025)",
            "time_duration": "Time & Duration (CET/CEST)",
            "company": "Company",
            "role": "Role/Purpose",
            "medium": "Meeting medium",
            "source": "Source(s)",
            "contacts": "Contacts (short)",
            "notes": "Additional notes",
        }
    },
    "interview_dataset_updated_2025.csv": {
        "canonical": {
            "role": "Role/Purpose",
            "company_contact": "Company/Contact",
            "notes": "Notes",
            "source": "Sources",
            "medium": "Medium",
            "contacts": "Contacts",
            "time_duration": "Time & Duration",
            "date": "Date (2025)",
        }
    },
}


# -----------------------------
# Core steps
# -----------------------------

def normalize_curated_df(curated_dir: Path, export_root: Optional[Path]) -> pd.DataFrame:
    p2022 = curated_dir / "interview_dataset_2022.csv"
    p2025 = curated_dir / "interview_dataset.csv"
    p2025u = curated_dir / "interview_dataset_updated_2025.csv"

    rows: List[Dict[str, object]] = []
    # 2022
    if p2022.exists():
        df = robust_read_csv(p2022)
        m = SCHEMA["interview_dataset_2022.csv"]["canonical"]
        for _, r in df.iterrows():
            rows.append({
                "event_date": r.get(m["date"]),
                "start_time": r.get(m["start_time"]),
                "end_time": r.get(m["end_time"]),
                "duration_minutes": None,
                "company": r.get(m["company"]),
                "role": r.get(m["role"]),
                "medium": r.get(m["medium"]),
                "source": r.get(m["source"]),
                "contacts": r.get(m["contacts"]),
                "notes": r.get(m["notes"]),
                "period": 2022,
            })
    # 2025 (original)
    if p2025.exists():
        df = robust_read_csv(p2025)
        m = SCHEMA["interview_dataset.csv"]["canonical"]
        for _, r in df.iterrows():
            st, en, dur = parse_time_range(r.get(m["time_duration"]))
            date_text = r.get(m["date"])
            dates = extract_dates_from_text(date_text)
            if dates:
                for d in dates:
                    rows.append({
                        "event_date": d,
                        "start_time": st,
                        "end_time": en,
                        "duration_minutes": dur,
                        "company": r.get(m["company"]),
                        "role": r.get(m["role"]),
                        "medium": r.get(m["medium"]),
                        "source": r.get(m["source"]),
                        "contacts": r.get(m["contacts"]),
                        "notes": r.get(m["notes"]),
                        "period": 2025,
                    })
            else:
                rows.append({
                    "event_date": date_text,
                    "start_time": st,
                    "end_time": en,
                    "duration_minutes": dur,
                    "company": r.get(m["company"]),
                    "role": r.get(m["role"]),
                    "medium": r.get(m["medium"]),
                    "source": r.get(m["source"]),
                    "contacts": r.get(m["contacts"]),
                    "notes": r.get(m["notes"]),
                    "period": 2025,
                })
    # 2025 (updated)
    if p2025u.exists():
        df = robust_read_csv(p2025u)
        m = SCHEMA["interview_dataset_updated_2025.csv"]["canonical"]
        for _, r in df.iterrows():
            st, en, dur = parse_time_range(r.get(m["time_duration"]))
            date_text = r.get(m["date"])
            dates = extract_dates_from_text(date_text)
            company_val = r.get(m["company_contact"]) or r.get(m["contacts"])
            if dates:
                for d in dates:
                    rows.append({
                        "event_date": d,
                        "start_time": st,
                        "end_time": en,
                        "duration_minutes": dur,
                        "company": company_val,
                        "role": r.get(m["role"]),
                        "medium": r.get(m["medium"]),
                        "source": r.get(m["source"]),
                        "contacts": r.get(m["contacts"]),
                        "notes": r.get(m["notes"]),
                        "period": 2025,
                    })
            else:
                rows.append({
                    "event_date": date_text,
                    "start_time": st,
                    "end_time": en,
                    "duration_minutes": dur,
                    "company": company_val,
                    "role": r.get(m["role"]),
                    "medium": r.get(m["medium"]),
                    "source": r.get(m["source"]),
                    "contacts": r.get(m["contacts"]),
                    "notes": r.get(m["notes"]),
                    "period": 2025,
                })

    norm = pd.DataFrame(rows)
    if not norm.empty:
        norm["event_date"] = pd.to_datetime(norm["event_date"], errors="coerce", dayfirst=False, infer_datetime_format=True)
        norm["year"] = norm["event_date"].dt.year
        for col in ("__source__", "__key__", "__login_time__"):
            if col in norm.columns:
                norm.drop(columns=[col], inplace=True)

    # Country inference from Logins.csv (optional)
    if export_root and (export_root / "Logins.csv").exists() and not norm.empty:
        logins = robust_read_csv(export_root / "Logins.csv")
        time_col = next((c for c in ["Date", "Login Time", "Time", "Timestamp", "Created At"] if c in logins.columns), None)
        country_col = next((c for c in ["Country", "country", "Location", "Country/Region", "Geo", "Region"] if c in logins.columns), None)
        logins["__time__"] = pd.to_datetime(logins[time_col], errors="coerce") if time_col else pd.NaT
        logins = logins.dropna(subset=["__time__"]).sort_values("__time__")
        logins["__country__"] = logins[country_col] if country_col else None
        tmp = norm[["event_date"]].copy().rename(columns={"event_date": "__time__"})
        tmp = tmp.dropna(subset=["__time__"]).sort_values("__time__")
        merged = pd.merge_asof(
            tmp,
            logins[["__time__", "__country__"]].rename(columns={"__time__": "__login_time__"}),
            left_on="__time__",
            right_on="__login_time__",
            tolerance=pd.Timedelta(days=5),
            direction="nearest",
        )
        norm = pd.concat([norm.reset_index(drop=True), merged[["__login_time__", "__country__"]]], axis=1)
        norm.rename(columns={"__country__": "inferred_country"}, inplace=True)

    return norm

def extract_linkedin_export(zip_dir: Path, op: Path) -> Optional[Path]:
    # Prefer latest Basic export; fallback to latest of any
    basic = list(zip_dir.glob("Basic_LinkedInDataExport*.zip*"))
    complete = list(zip_dir.glob("Complete_LinkedInDataExport*.zip*"))
    pick_from = basic if basic else (basic + complete)
    if not pick_from:
        pick_from = complete
    if not pick_from:
        return None
    z = sorted(pick_from, key=lambda p: p.stat().st_mtime, reverse=True)[0]
    dest = op / "input" / "linkedin_export_2025"
    dest.mkdir(parents=True, exist_ok=True)
    try:
        with zipfile.ZipFile(z, "r") as zf:
            zf.extractall(dest)
    except zipfile.BadZipFile:
        # Some exports are double-zipped or have spaces. Try renaming fallback
        try:
            with zipfile.ZipFile(str(z), "r") as zf:
                zf.extractall(dest)
        except Exception:
            return None
    return dest


def inventory_export(export_root: Path, processing_dir: Path, output_dir: Path, minimal: bool = False) -> Dict[str, Dict[str, object]]:
    catalog: Dict[str, Dict[str, object]] = {}
    if export_root and export_root.exists():
        for p in export_root.rglob("*.csv"):
            rel = str(p.relative_to(export_root))
            try:
                with open(p, encoding="utf-8-sig", newline="") as f:
                    r = csv.reader(f)
                    header = next(r, [])
                    n = 0
                    for _ in r:
                        n += 1
                catalog[rel] = {"rows": n, "columns": header}
            except Exception as e:
                catalog[rel] = {"error": str(e)}
    (processing_dir / "catalog.json").write_text(json.dumps(catalog, indent=2), encoding="utf-8")

    if not minimal:
        # Convenience: export key CSVs to Excel for browsing
        key_targets = {
            "messages.csv": ["messages.csv"],
            "Invitations.csv": ["Invitations.csv"],
            "Logins.csv": ["Logins.csv"],
            "Profile.csv": ["Profile.csv"],
        }
        (output_dir / "raw_exports").mkdir(parents=True, exist_ok=True)
        for name, alts in key_targets.items():
            src = None
            for a in alts:
                p = export_root / a if export_root else None
                if p and p.exists():
                    src = p
                    break
            if not src:
                continue
            df = robust_read_csv(src)
            # write with openpyxl to avoid auto URL hyperlinking issues
            out_xlsx = output_dir / "raw_exports" / (Path(name).stem + ".xlsx")
            with pd.ExcelWriter(out_xlsx, engine="openpyxl") as w:
                df.to_excel(w, index=False, sheet_name="Sheet1")
        # Dataset catalog workbook
        rows = []
        for rel, info in catalog.items():
            row = {"file": rel}
            if isinstance(info, dict):
                row["rows"] = info.get("rows")
                cols = info.get("columns")
                row["columns"] = ", ".join(cols) if cols else None
                row["error"] = info.get("error")
            rows.append(row)
        # Always use openpyxl for consistent writing
        cat_path = output_dir / "dataset_catalog.xlsx"
        with pd.ExcelWriter(cat_path, engine="openpyxl") as w:
            pd.DataFrame(rows).sort_values("file").to_excel(w, index=False, sheet_name="Sheet1")
    return catalog


def normalize_curated(curated_dir: Path, export_root: Optional[Path], op: Path) -> Path:
    p2022 = curated_dir / "interview_dataset_2022.csv"
    p2025 = curated_dir / "interview_dataset.csv"
    p2025u = curated_dir / "interview_dataset_updated_2025.csv"

    rows: List[Dict[str, object]] = []
    # 2022
    if p2022.exists():
        df = robust_read_csv(p2022)
        m = SCHEMA["interview_dataset_2022.csv"]["canonical"]
        for _, r in df.iterrows():
            rows.append({
                "event_date": r.get(m["date"]),
                "start_time": r.get(m["start_time"]),
                "end_time": r.get(m["end_time"]),
                "duration_minutes": None,
                "company": r.get(m["company"]),
                "role": r.get(m["role"]),
                "medium": r.get(m["medium"]),
                "source": r.get(m["source"]),
                "contacts": r.get(m["contacts"]),
                "notes": r.get(m["notes"]),
                "period": 2022,
            })
    # 2025 (original)
    if p2025.exists():
        df = robust_read_csv(p2025)
        m = SCHEMA["interview_dataset.csv"]["canonical"]
        for _, r in df.iterrows():
            st, en, dur = parse_time_range(r.get(m["time_duration"]))
            date_text = r.get(m["date"])
            dates = extract_dates_from_text(date_text)
            if dates:
                for d in dates:
                    rows.append({
                        "event_date": d,
                        "start_time": st,
                        "end_time": en,
                        "duration_minutes": dur,
                        "company": r.get(m["company"]),
                        "role": r.get(m["role"]),
                        "medium": r.get(m["medium"]),
                        "source": r.get(m["source"]),
                        "contacts": r.get(m["contacts"]),
                        "notes": r.get(m["notes"]),
                        "period": 2025,
                    })
            else:
                rows.append({
                    "event_date": date_text,
                    "start_time": st,
                    "end_time": en,
                    "duration_minutes": dur,
                    "company": r.get(m["company"]),
                    "role": r.get(m["role"]),
                    "medium": r.get(m["medium"]),
                    "source": r.get(m["source"]),
                    "contacts": r.get(m["contacts"]),
                    "notes": r.get(m["notes"]),
                    "period": 2025,
                })
    # 2025 (updated)
    if p2025u.exists():
        df = robust_read_csv(p2025u)
        m = SCHEMA["interview_dataset_updated_2025.csv"]["canonical"]
        for _, r in df.iterrows():
            st, en, dur = parse_time_range(r.get(m["time_duration"]))
            date_text = r.get(m["date"])
            dates = extract_dates_from_text(date_text)
            company_val = r.get(m["company_contact"]) or r.get(m["contacts"])
            if dates:
                for d in dates:
                    rows.append({
                        "event_date": d,
                        "start_time": st,
                        "end_time": en,
                        "duration_minutes": dur,
                        "company": company_val,
                        "role": r.get(m["role"]),
                        "medium": r.get(m["medium"]),
                        "source": r.get(m["source"]),
                        "contacts": r.get(m["contacts"]),
                        "notes": r.get(m["notes"]),
                        "period": 2025,
                    })
            else:
                rows.append({
                    "event_date": date_text,
                    "start_time": st,
                    "end_time": en,
                    "duration_minutes": dur,
                    "company": company_val,
                    "role": r.get(m["role"]),
                    "medium": r.get(m["medium"]),
                    "source": r.get(m["source"]),
                    "contacts": r.get(m["contacts"]),
                    "notes": r.get(m["notes"]),
                    "period": 2025,
                })

    norm = pd.DataFrame(rows)
    if not norm.empty:
        norm["event_date"] = pd.to_datetime(norm["event_date"], errors="coerce", dayfirst=False, infer_datetime_format=True)
        norm["year"] = norm["event_date"].dt.year
        # Ensure internal columns are not present
        for col in ("__source__", "__key__", "__login_time__"):
            if col in norm.columns:
                norm.drop(columns=[col], inplace=True)

    # Country inference from Logins.csv (optional)
    if export_root and (export_root / "Logins.csv").exists() and not norm.empty:
        logins = robust_read_csv(export_root / "Logins.csv")
        time_col = next((c for c in ["Date", "Login Time", "Time", "Timestamp", "Created At"] if c in logins.columns), None)
        country_col = next((c for c in ["Country", "country", "Location", "Country/Region", "Geo", "Region"] if c in logins.columns), None)
        logins["__time__"] = pd.to_datetime(logins[time_col], errors="coerce") if time_col else pd.NaT
        logins = logins.dropna(subset=["__time__"]).sort_values("__time__")
        logins["__country__"] = logins[country_col] if country_col else None
        tmp = norm[["event_date"]].copy().rename(columns={"event_date": "__time__"})
        tmp = tmp.dropna(subset=["__time__"]).sort_values("__time__")
        merged = pd.merge_asof(
            tmp,
            logins[["__time__", "__country__"]].rename(columns={"__time__": "__login_time__"}),
            left_on="__time__",
            right_on="__login_time__",
            tolerance=pd.Timedelta(days=5),
            direction="nearest",
        )
        norm = pd.concat([norm.reset_index(drop=True), merged[["__login_time__", "__country__"]]], axis=1)
        norm.rename(columns={"__country__": "inferred_country"}, inplace=True)

    out_path = op / "output" / "enriched" / "normalized_interviews.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        norm.to_excel(w, index=False, sheet_name="Sheet1")
    try:
        wb = load_workbook(str(out_path))
        if "PROVENANCE" in wb.sheetnames:
            wb.remove(wb["PROVENANCE"])
        # Append PROVENANCE as last sheet (keep data sheets first)
        ws = wb.create_sheet("PROVENANCE")
        ws["A1"] = "operation_id"; ws["B1"] = op.name
        ws["A2"] = "generated_at_utc"; ws["B2"] = datetime.utcnow().isoformat() + "Z"
        ws["A3"] = "sources"; ws["B3"] = ", ".join([
            str(p2022) if p2022.exists() else "",
            str(p2025) if p2025.exists() else "",
            str(p2025u) if p2025u.exists() else "",
            str(export_root / "Logins.csv") if export_root and (export_root / "Logins.csv").exists() else "",
        ])
        ws["A4"] = "output_path"; ws["B4"] = str(out_path)
        wb.save(str(out_path))
    except Exception:
        pass
    # QA summary
    qa = {
        "rows_total": int(len(norm)),
        "rows_with_country": int(norm["inferred_country"].notna().sum()) if "inferred_country" in norm.columns else 0,
        "rows_missing_date": int(norm["event_date"].isna().sum()) if "event_date" in norm.columns else 0,
    }
    (op / "output" / "report" / "join_qa.json").write_text(json.dumps(qa, indent=2), encoding="utf-8")
    return out_path


def build_linkedin_events(export_root: Optional[Path], op: Path) -> Optional[Path]:
    if not export_root or not export_root.exists():
        return None
    msgs_p = export_root / "messages.csv"
    inv_p = export_root / "Invitations.csv"
    profile_p = export_root / "Profile.csv"
    msgs = robust_read_csv(msgs_p) if msgs_p.exists() else pd.DataFrame()
    inv = robust_read_csv(inv_p) if inv_p.exists() else pd.DataFrame()
    profile = robust_read_csv(profile_p) if profile_p.exists() else pd.DataFrame()
    user_fullname = None
    try:
        if not profile.empty:
            fn = next((c for c in ["First Name","FirstName","first_name"] if c in profile.columns), None)
            ln = next((c for c in ["Last Name","LastName","last_name"] if c in profile.columns), None)
            if fn and ln:
                user_fullname = f"{str(profile.iloc[0][fn]).strip()} {str(profile.iloc[0][ln]).strip()}".strip()
    except Exception:
        user_fullname = None

    kw_en = [
        r"\binterview\b", r"\bscreen(ing)?\b", r"\b(call|chat|conversation)\b", r"\bschedul(e|ing)\b",
        r"\bavailability\b", r"\bslot(s)?\b", r"\bbook\b", r"\binvit(e|ation)\b", r"\bzoom\b", r"\bteams\b",
        r"\bgoogle\s*meet\b", r"\bcalendly\b", r"\bcalendar\b", r"\b(first|1st)\s*round\b", r"\b(second|2nd)\s*round\b",
        r"\bstage\b", r"\bassessment\b", r"\btest\b", r"\btalent\s*partner\b", r"\brecruiter\b", r"\bhiring\s*manager\b",
    ]
    kw_fr = [
        r"\bentretien(s)?\b", r"\brdv\b", r"\brendez[- ]vous\b", r"\bappel(s)?\b", r"\bt[ée]l[ée]phonique\b",
        r"\bvisioconf[ée]rence\b", r"\br[ée]union\b", r"\bdisponibilit[ée]s?\b", r"\bcr[ée]neau(x)?\b",
        r"\bplanifier\b", r"\bplanification\b", r"\binvit(ation|er)\b", r"\bcalendly\b", r"\bcalendrier\b",
        r"\bpremier\s*tour\b", r"\bdeuxi[èe]me\s*tour\b", r"\b2e\s*tour\b", r"\btest\b", r"\b[ée]valuation\b",
        r"\b[ée]tude de cas\b", r"\brecruteur\b", r"\bcharg[ée] de recrutement\b",
    ]
    kw_re = re.compile("|".join(kw_en + kw_fr), flags=re.IGNORECASE)

    rows: List[Dict[str, object]] = []
    if not msgs.empty:
        # parse time
        t = None
        for c in ["Date", "DATE", "Sent At", "Created At", "Timestamp"]:
            if c in msgs.columns:
                try:
                    t = pd.to_datetime(msgs[c], errors="coerce")
                    break
                except Exception:
                    continue
        if t is None:
            for c in msgs.columns:
                try:
                    tt = pd.to_datetime(msgs[c], errors="coerce")
                    if tt.notna().sum() > max(3, int(0.1 * len(tt))):
                        t = tt
                        break
                except Exception:
                    continue
        msgs = msgs.assign(__time__=t)
        msgs = msgs.dropna(subset=["__time__"]).sort_values("__time__")
        from_col = next((c for c in ["From", "FROM", "Sender", "Author"] if c in msgs.columns), None)
        to_col = next((c for c in ["To", "TO", "Recipient"] if c in msgs.columns), None)
        subj_col = next((c for c in ["Subject", "SUBJECT", "subject"] if c in msgs.columns), None)
        body_col = next((c for c in ["Content", "Body", "Message", "Text"] if c in msgs.columns), None)
        for _, r in msgs.iterrows():
            text = " ".join([str(r.get(subj_col) or ""), str(r.get(body_col) or "")]) if subj_col or body_col else ""
            if kw_re.search(text):
                from_name = str(r.get(from_col) or "").strip()
                to_name = str(r.get(to_col) or "").strip()
                direction = "outbound" if user_fullname and from_name and from_name.lower() == user_fullname.lower() else "inbound"
                counterparty = to_name if direction == "outbound" else from_name
                rows.append({
                    "event_date": r["__time__"].date().isoformat(),
                    "start_time": None,
                    "end_time": None,
                    "duration_minutes": None,
                    "company": counterparty,
                    "role": None,
                    "medium": "LinkedIn DM",
                    "source": "LinkedIn messages",
                    "contacts": counterparty,
                    "notes": (text[:300]) if text else None,
                    "period": int(r["__time__"].year),
                    "direction": direction,
                })

    if not inv.empty:
        t = None
        for c in ["Date", "Sent At", "Created At", "Timestamp"]:
            if c in inv.columns:
                try:
                    t = pd.to_datetime(inv[c], errors="coerce")
                    break
                except Exception:
                    continue
        inv = inv.assign(__time__=t)
        inv = inv.dropna(subset=["__time__"]).sort_values("__time__")
        from_col = next((c for c in ["From", "Invite From", "Sender", "Author"] if c in inv.columns), None)
        for _, r in inv.iterrows():
            from_name = str(r.get(from_col) or "").strip()
            direction = "inbound" if user_fullname and from_name.lower() != user_fullname.lower() else "outbound"
            rows.append({
                "event_date": r["__time__"].date().isoformat(),
                "start_time": None,
                "end_time": None,
                "duration_minutes": None,
                "company": from_name,
                "role": None,
                "medium": "LinkedIn Invitation",
                "source": "LinkedIn invitations",
                "contacts": from_name,
                "notes": "LinkedIn invitation",
                "period": int(r["__time__"].year),
                "direction": direction,
            })

    if not rows:
        return None

    lk = pd.DataFrame(rows)
    out_path = op / "output" / "enriched" / "linkedin_events.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        lk.to_excel(w, index=False, sheet_name="Sheet1")
    try:
        wb = load_workbook(str(out_path))
        if "PROVENANCE" in wb.sheetnames:
            wb.remove(wb["PROVENANCE"])
        # Append PROVENANCE as last sheet
        ws = wb.create_sheet("PROVENANCE")
        ws["A1"] = "operation_id"; ws["B1"] = op.name
        ws["A2"] = "generated_at_utc"; ws["B2"] = datetime.utcnow().isoformat() + "Z"
        ws["A3"] = "sources"; ws["B3"] = "messages.csv, Invitations.csv"
        ws["A4"] = "output_path"; ws["B4"] = str(out_path)
        wb.save(str(out_path))
    except Exception:
        pass
    return out_path


def merge_events(op: Path, curated_path: Path, linkedin_path: Optional[Path]) -> Path:
    curated = pd.read_excel(curated_path) if curated_path.exists() else pd.DataFrame()
    li = pd.read_excel(linkedin_path) if linkedin_path and linkedin_path.exists() else pd.DataFrame()
    if not curated.empty:
        curated["__source__"] = "curated"
        curated["event_type"] = "interview"
    if not li.empty:
        li["__source__"] = "linkedin"
        # Classify LinkedIn-derived events as outreach by default
        li["event_type"] = li.get("medium").fillna("").apply(lambda m: "outreach" if "LinkedIn" in str(m) else "outreach")
    cols = [
        "event_date",
        "start_time",
        "end_time",
        "duration_minutes",
        "company",
        "role",
        "medium",
        "source",
        "contacts",
        "notes",
        "period",
        "direction",
    ]
    for df in (curated, li):
        for c in cols:
            if c not in df.columns:
                df[c] = None
    # Keep event_type for downstream filters
    extra_cols = ["event_type"] if "event_type" in curated.columns or "event_type" in li.columns else []
    all_df = pd.concat([
        curated[cols + extra_cols + ["__source__"]],
        li[cols + extra_cols + ["__source__"]]
    ], ignore_index=True)
    all_df["event_date"] = pd.to_datetime(all_df["event_date"], errors="coerce")
    # Backfill missing core fields before dedupe
    if "role" in all_df.columns and "notes" in all_df.columns:
        all_df["role"] = all_df.apply(
            lambda r: r["role"] if pd.notna(r.get("role")) and str(r.get("role")).strip()
            else infer_role_from_text(r.get("notes")), axis=1
        )
    if "company" in all_df.columns and "contacts" in all_df.columns:
        all_df["company"] = all_df.apply(
            lambda r: r["company"] if pd.notna(r.get("company")) and str(r.get("company")).strip()
            else (r.get("contacts") if pd.notna(r.get("contacts")) and str(r.get("contacts")).strip() else None), axis=1
        )
        all_df["contacts"] = all_df.apply(
            lambda r: r["contacts"] if pd.notna(r.get("contacts")) and str(r.get("contacts")).strip()
            else (r.get("company") if pd.notna(r.get("company")) and str(r.get("company")).strip() else None), axis=1
        )
    key = all_df[["event_date", "medium", "company", "role"]].astype(str)
    all_df["__key__"] = key.apply(lambda r: "|".join(r.values), axis=1)
    merged = all_df.drop_duplicates(subset=["__key__"]).copy()
    # Remove internal dedupe key from exported data for cleanliness
    if "__key__" in merged.columns:
        merged.drop(columns=["__key__"], inplace=True)

    out_path = op / "output" / "enriched" / "merged_interview_events.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        merged.to_excel(w, index=False, sheet_name="Sheet1")
    try:
        wb = load_workbook(str(out_path))
        if "PROVENANCE" in wb.sheetnames:
            wb.remove(wb["PROVENANCE"])
        # Append PROVENANCE as last sheet
        ws = wb.create_sheet("PROVENANCE")
        ws["A1"] = "operation_id"; ws["B1"] = op.name
        ws["A2"] = "generated_at_utc"; ws["B2"] = datetime.utcnow().isoformat() + "Z"
        ws["A3"] = "sources"; ws["B3"] = f"{curated_path}, {linkedin_path}"
        ws["A4"] = "output_path"; ws["B4"] = str(out_path)
        wb.save(str(out_path))
    except Exception:
        pass
    return out_path


def build_linkedin_events_df(export_root: Optional[Path]) -> pd.DataFrame:
    if not export_root or not export_root.exists():
        return pd.DataFrame()
    msgs_p = export_root / "messages.csv"
    inv_p = export_root / "Invitations.csv"
    profile_p = export_root / "Profile.csv"
    msgs = robust_read_csv(msgs_p) if msgs_p.exists() else pd.DataFrame()
    inv = robust_read_csv(inv_p) if inv_p.exists() else pd.DataFrame()
    profile = robust_read_csv(profile_p) if profile_p.exists() else pd.DataFrame()
    user_fullname = None
    try:
        if not profile.empty:
            fn = next((c for c in ["First Name","FirstName","first_name"] if c in profile.columns), None)
            ln = next((c for c in ["Last Name","LastName","last_name"] if c in profile.columns), None)
            if fn and ln:
                user_fullname = f"{str(profile.iloc[0][fn]).strip()} {str(profile.iloc[0][ln]).strip()}".strip()
    except Exception:
        user_fullname = None

    kw_en = [
        r"\binterview\b", r"\bscreen(ing)?\b", r"\b(call|chat|conversation)\b", r"\bschedul(e|ing)\b",
        r"\bavailability\b", r"\bslot(s)?\b", r"\bbook\b", r"\binvit(e|ation)\b", r"\bzoom\b", r"\bteams\b",
        r"\bgoogle\s*meet\b", r"\bcalendly\b", r"\bcalendar\b", r"\b(first|1st)\s*round\b", r"\b(second|2nd)\s*round\b",
        r"\bstage\b", r"\bassessment\b", r"\btest\b", r"\btalent\s*partner\b", r"\brecruiter\b", r"\bhiring\s*manager\b",
    ]
    kw_fr = [
        r"\bentretien(s)?\b", r"\brdv\b", r"\brendez[- ]vous\b", r"\bappel(s)?\b", r"\bt[ée]l[ée]phonique\b",
        r"\bvisioconf[ée]rence\b", r"\br[ée]union\b", r"\bdisponibilit[ée]s?\b", r"\bcr[ée]neau(x)?\b",
        r"\bplanifier\b", r"\bplanification\b", r"\binvit(ation|er)\b", r"\bcalendly\b", r"\bcalendrier\b",
        r"\bpremier\s*tour\b", r"\bdeuxi[èe]me\s*tour\b", r"\b2e\s*tour\b", r"\btest\b", r"\b[ée]valuation\b",
        r"\b[ée]tude de cas\b", r"\brecruteur\b", r"\bcharg[ée] de recrutement\b",
    ]
    kw_re = re.compile("|".join(kw_en + kw_fr), flags=re.IGNORECASE)

    rows: List[Dict[str, object]] = []
    if not msgs.empty:
        t = None
        for c in ["Date", "DATE", "Sent At", "Created At", "Timestamp"]:
            if c in msgs.columns:
                try:
                    t = pd.to_datetime(msgs[c], errors="coerce")
                    break
                except Exception:
                    continue
        if t is None:
            for c in msgs.columns:
                try:
                    tt = pd.to_datetime(msgs[c], errors="coerce")
                    if tt.notna().sum() > max(3, int(0.1 * len(tt))):
                        t = tt
                        break
                except Exception:
                    continue
        msgs = msgs.assign(__time__=t)
        msgs = msgs.dropna(subset=["__time__"]).sort_values("__time__")
        from_col = next((c for c in ["From", "FROM", "Sender", "Author"] if c in msgs.columns), None)
        to_col = next((c for c in ["To", "TO", "Recipient"] if c in msgs.columns), None)
        subj_col = next((c for c in ["Subject", "SUBJECT", "subject"] if c in msgs.columns), None)
        body_col = next((c for c in ["Content", "Body", "Message", "Text"] if c in msgs.columns), None)
        for _, r in msgs.iterrows():
            text = " ".join([str(r.get(subj_col) or ""), str(r.get(body_col) or "")]) if subj_col or body_col else ""
            if kw_re.search(text):
                from_name = str(r.get(from_col) or "").strip()
                to_name = str(r.get(to_col) or "").strip()
                direction = "outbound" if user_fullname and from_name and from_name.lower() == user_fullname.lower() else "inbound"
                counterparty = to_name if direction == "outbound" else from_name
                rows.append({
                    "event_date": r["__time__"].date().isoformat(),
                    "start_time": None,
                    "end_time": None,
                    "duration_minutes": None,
                    "company": counterparty,
                    "role": None,
                    "medium": "LinkedIn DM",
                    "source": "LinkedIn messages",
                    "contacts": counterparty,
                    "notes": (text[:300]) if text else None,
                    "period": int(r["__time__"].year),
                    "direction": direction,
                })

    if not inv.empty:
        t = None
        for c in ["Date", "Sent At", "Created At", "Timestamp"]:
            if c in inv.columns:
                try:
                    t = pd.to_datetime(inv[c], errors="coerce")
                    break
                except Exception:
                    continue
        inv = inv.assign(__time__=t)
        inv = inv.dropna(subset=["__time__"]).sort_values("__time__")
        from_col = next((c for c in ["From", "Invite From", "Sender", "Author"] if c in inv.columns), None)
        for _, r in inv.iterrows():
            from_name = str(r.get(from_col) or "").strip()
            direction = "inbound" if user_fullname and from_name.lower() != user_fullname.lower() else "outbound"
            rows.append({
                "event_date": r["__time__"].date().isoformat(),
                "start_time": None,
                "end_time": None,
                "duration_minutes": None,
                "company": from_name,
                "role": None,
                "medium": "LinkedIn Invitation",
                "source": "LinkedIn invitations",
                "contacts": from_name,
                "notes": "LinkedIn invitation",
                "period": int(r["__time__"].year),
                "direction": direction,
            })

    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows)


def merge_events_df(curated_df: pd.DataFrame, li_df: pd.DataFrame) -> pd.DataFrame:
    curated = curated_df.copy() if curated_df is not None else pd.DataFrame()
    li = li_df.copy() if li_df is not None else pd.DataFrame()
    if not curated.empty:
        curated["__source__"] = "curated"
        curated["event_type"] = "interview"
    if not li.empty:
        li["__source__"] = "linkedin"
        li["event_type"] = li.get("medium").fillna("").apply(lambda m: "outreach" if "LinkedIn" in str(m) else "outreach")
    cols = [
        "event_date",
        "start_time",
        "end_time",
        "duration_minutes",
        "company",
        "role",
        "medium",
        "source",
        "contacts",
        "notes",
        "period",
        "direction",
    ]
    for df in (curated, li):
        for c in cols:
            if c not in df.columns:
                df[c] = None
    extra_cols = ["event_type"] if ("event_type" in curated.columns or "event_type" in li.columns) else []
    all_df = pd.concat([
        curated[cols + extra_cols + ["__source__"]],
        li[cols + extra_cols + ["__source__"]]
    ], ignore_index=True)
    all_df["event_date"] = pd.to_datetime(all_df["event_date"], errors="coerce")
    if "role" in all_df.columns and "notes" in all_df.columns:
        all_df["role"] = all_df.apply(
            lambda r: r["role"] if pd.notna(r.get("role")) and str(r.get("role")).strip()
            else infer_role_from_text(r.get("notes")), axis=1
        )
    if "company" in all_df.columns and "contacts" in all_df.columns:
        all_df["company"] = all_df.apply(
            lambda r: r["company"] if pd.notna(r.get("company")) and str(r.get("company")).strip()
            else (r.get("contacts") if pd.notna(r.get("contacts")) and str(r.get("contacts")).strip() else None), axis=1
        )
        all_df["contacts"] = all_df.apply(
            lambda r: r["contacts"] if pd.notna(r.get("contacts")) and str(r.get("contacts")).strip()
            else (r.get("company") if pd.notna(r.get("company")) and str(r.get("company")).strip() else None), axis=1
        )
    key = all_df[["event_date", "medium", "company", "role"]].astype(str)
    all_df["__key__"] = key.apply(lambda r: "|".join(r.values), axis=1)
    merged = all_df.drop_duplicates(subset=["__key__"]).copy()
    if "__key__" in merged.columns:
        merged.drop(columns=["__key__"], inplace=True)
    return merged


def extract_applications_events_df(export_root: Optional[Path], curated_dir: Optional[Path] = None) -> Tuple[pd.DataFrame, Dict[str, object]]:
    if not export_root or not export_root.exists():
        return pd.DataFrame(), {"status": "skipped", "reason": "no_export"}

    jobs_dirs = [d for d in export_root.rglob("*") if d.is_dir() and d.name.lower() == "jobs"]
    candidate_files: List[Path] = []
    if jobs_dirs:
        for jd in jobs_dirs:
            candidate_files.extend(list(jd.rglob("*.csv")))
    else:
        candidate_files = list(export_root.rglob("*.csv"))

    def detect_date_column(df: pd.DataFrame) -> Optional[str]:
        lower = {c.lower(): c for c in df.columns}
        candidates = [
            "applied on",
            "appliedon",
            "application date",
            "date applied",
            "applied at",
            "date",
            "created at",
            "timestamp",
        ]
        for k in candidates:
            if k in lower:
                return lower[k]
        for c in df.columns:
            try:
                pd.to_datetime(df[c])
                return c
            except Exception:
                continue
        return None

    apps_frames: List[pd.DataFrame] = []
    picked_files: List[str] = []
    for p in candidate_files:
        try:
            df = robust_read_csv(p)
        except Exception:
            continue
        date_col = detect_date_column(df)
        if not date_col:
            continue
        cols_lower = {c.lower(): c for c in df.columns}
        has_company = any(k in cols_lower for k in ["company", "company name", "organization", "employer"])
        has_title = any(k in cols_lower for k in ["job title", "position", "title", "role", "job"])
        if not (has_company or has_title):
            continue
        tmp = df.copy()
        tmp["__date__"] = pd.to_datetime(tmp[date_col], errors="coerce")
        tmp = tmp.dropna(subset=["__date__"]).sort_values("__date__")
        if tmp.empty:
            continue
        lower = {c.lower(): c for c in tmp.columns}
        def col(*names: str) -> Optional[str]:
            for n in names:
                if n in lower:
                    return lower[n]
            return None
        company_c = col("company","company name","organization","employer")
        title_c = col("job title","position","title","role","job")
        link_c = col("job url","joburl","link","url")
        location_c = col("location","city","country")
        sel = {
            "applied_date": tmp[date_col],
            "__date__": tmp["__date__"],
            "company": tmp[company_c] if company_c else None,
            "title": tmp[title_c] if title_c else None,
            "location": tmp[location_c] if location_c else None,
            "link": tmp[link_c] if link_c else None,
            "source_file": str(p)
        }
        df_sel = pd.DataFrame(sel)
        apps_frames.append(df_sel)
        picked_files.append(str(p))

    if not apps_frames and curated_dir:
        for name in ["job_applications_2025_sample.csv", "job_applications_2025_monthly_counts.csv"]:
            p = curated_dir / name
            if p.exists():
                try:
                    df = robust_read_csv(p)
                    date_col = detect_date_column(df)
                    if date_col:
                        df["__date__"] = pd.to_datetime(df[date_col], errors="coerce")
                        df = df.dropna(subset=["__date__"]).sort_values("__date__")
                        if not df.empty:
                            apps_frames.append(df[[date_col]].rename(columns={date_col: "applied_date"}).assign(__date__=df["__date__"]))
                            picked_files.append(str(p))
                except Exception:
                    continue

    if not apps_frames:
        return pd.DataFrame(), {"status": "skipped", "reason": "no_applications_table_found"}

    apps = pd.concat(apps_frames, ignore_index=True)
    if apps.empty:
        return pd.DataFrame(), {"status": "skipped", "reason": "no_valid_dates"}

    apps["year"] = apps["__date__"].dt.year
    by_year = apps.groupby("year").size().rename("applications")

    dates = apps["__date__"].tolist()
    dates_sorted = sorted(dates)
    periods: List[List[datetime]] = []
    if dates_sorted:
        cur = [dates_sorted[0]]
        for d in dates_sorted[1:]:
            if (d - cur[-1]).days > 7:
                periods.append(cur)
                cur = [d]
            else:
                cur.append(d)
        periods.append(cur)

    period_rows = []
    for i, grp in enumerate(periods, start=1):
        start = grp[0]
        end = grp[-1]
        num_days = max(1, (end - start).days + 1)
        num_apps = len(grp)
        period_rows.append({
            "period_index": i,
            "start_date": start.date().isoformat(),
            "end_date": end.date().isoformat(),
            "num_days": num_days,
            "num_applications": num_apps,
            "applications_per_day": round(num_apps / num_days, 4),
        })

    summary = {
        "status": "success",
        "input_files": picked_files,
        "applications_by_year": by_year.to_dict(),
        "periods": period_rows,
    }
    return apps, summary


def export_minimal_deliverables(op: Path, merged_df: pd.DataFrame, apps_df: pd.DataFrame) -> Dict[str, str]:
    enriched_dir = op / "output" / "enriched"
    enriched_dir.mkdir(parents=True, exist_ok=True)

    # Interviews dataset (events: interview + outreach)
    flat = merged_df.copy() if merged_df is not None else pd.DataFrame()
    for c in ["event_year","event_type","__source__"]:
        if c not in flat.columns:
            flat[c] = None
    flat["event_datetime"] = pd.to_datetime(flat.get("event_date"), errors="coerce")
    fallback_year = pd.to_numeric(flat.get("period"), errors="coerce")
    flat["event_year"] = pd.to_numeric(flat.get("event_year"), errors="coerce")
    mask_year = flat["event_year"].isna() & fallback_year.notna()
    if mask_year.any():
        flat.loc[mask_year, "event_year"] = fallback_year[mask_year].astype(int)
    mask_dt = flat["event_datetime"].isna() & fallback_year.notna()
    if mask_dt.any():
        flat.loc[mask_dt, "event_datetime"] = pd.to_datetime(
            fallback_year[mask_dt].astype(int).astype(str) + "-01-01",
            errors="coerce"
        )
    flat_out = flat[[
        "event_datetime","event_year","event_type","company","contacts","role","medium","source","notes","period","start_time","end_time","duration_minutes","__source__"
    ]].copy()
    try:
        flat_out.sort_values(["event_year","event_datetime","company"], inplace=True)
    except Exception:
        pass
    interviews_path = enriched_dir / "interviews_dataset.xlsx"
    with pd.ExcelWriter(interviews_path, engine="openpyxl") as w:
        flat_out.to_excel(w, index=False, sheet_name="Events")
    try:
        wb = load_workbook(str(interviews_path))
        if "PROVENANCE" in wb.sheetnames:
            wb.remove(wb["PROVENANCE"])
        ws = wb.create_sheet("PROVENANCE")
        ws["A1"] = "operation_id"; ws["B1"] = op.name
        ws["A2"] = "generated_at_utc"; ws["B2"] = datetime.utcnow().isoformat() + "Z"
        ws["A3"] = "output_path"; ws["B3"] = str(interviews_path)
        wb.save(str(interviews_path))
    except Exception:
        pass

    # Applications dataset
    apps = apps_df.copy() if apps_df is not None else pd.DataFrame()
    apps_out_cols = ["applied_date","company","title","location","link","year","source_file"]
    for c in apps_out_cols:
        if c not in apps.columns:
            apps[c] = None
    apps_path = enriched_dir / "applications_dataset.xlsx"
    with pd.ExcelWriter(apps_path, engine="openpyxl") as w:
        apps[apps_out_cols].to_excel(w, index=False, sheet_name="Applications")
        # Optional: 2022/2025 sheets if present
        for yr in (2022, 2025):
            sub = apps[apps.get("year") == yr]
            if not sub.empty:
                sub[apps_out_cols].to_excel(w, index=False, sheet_name=f"Applications_{yr}")
    try:
        wb = load_workbook(str(apps_path))
        if "PROVENANCE" in wb.sheetnames:
            wb.remove(wb["PROVENANCE"])
        ws = wb.create_sheet("PROVENANCE")
        ws["A1"] = "operation_id"; ws["B1"] = op.name
        ws["A2"] = "generated_at_utc"; ws["B2"] = datetime.utcnow().isoformat() + "Z"
        ws["A3"] = "output_path"; ws["B3"] = str(apps_path)
        wb.save(str(apps_path))
    except Exception:
        pass

    return {"interviews_dataset": str(interviews_path), "applications_dataset": str(apps_path)}

def detect_and_analyze_applications(export_root: Optional[Path], op: Path, curated_dir: Optional[Path] = None) -> Dict[str, object]:
    summary: Dict[str, object] = {"status": "skipped", "reason": "no_export"}
    if not export_root or not export_root.exists():
        return summary

    # Look specifically for a jobs subfolder inside the export and combine CSVs
    jobs_dirs = [d for d in export_root.rglob("*") if d.is_dir() and d.name.lower() == "jobs"]
    candidate_files: List[Path] = []
    if jobs_dirs:
        for jd in jobs_dirs:
            candidate_files.extend(list(jd.rglob("*.csv")))
    else:
        # fallback: scan all CSVs
        candidate_files = list(export_root.rglob("*.csv"))

    def detect_date_column(df: pd.DataFrame) -> Optional[str]:
        lower = {c.lower(): c for c in df.columns}
        candidates = [
            "applied on",
            "appliedon",
            "application date",
            "date applied",
            "applied at",
            "date",
            "created at",
            "timestamp",
        ]
        for k in candidates:
            if k in lower:
                return lower[k]
        for c in df.columns:
            try:
                pd.to_datetime(df[c])
                return c
            except Exception:
                continue
        return None

    apps_frames: List[pd.DataFrame] = []
    picked_files: List[str] = []
    for p in candidate_files:
        try:
            df = robust_read_csv(p)
        except Exception:
            continue
        date_col = detect_date_column(df)
        if not date_col:
            continue
        # Require either company or title-like column to avoid noise tables
        cols_lower = {c.lower(): c for c in df.columns}
        has_company = any(k in cols_lower for k in ["company", "company name", "organization", "employer"])
        has_title = any(k in cols_lower for k in ["job title", "position", "title", "role", "job"]) 
        if not (has_company or has_title):
            continue
        tmp = df.copy()
        tmp["__date__"] = pd.to_datetime(tmp[date_col], errors="coerce")
        tmp = tmp.dropna(subset=["__date__"]).sort_values("__date__")
        if tmp.empty:
            continue
        # Map helpful columns where available
        lower = {c.lower(): c for c in tmp.columns}
        def col(*names: str) -> Optional[str]:
            for n in names:
                if n in lower:
                    return lower[n]
            return None
        company_c = col("company","company name","organization","employer")
        title_c = col("job title","position","title","role","job")
        link_c = col("job url","joburl","link","url")
        location_c = col("location","city","country")
        sel = {
            "applied_date": tmp[date_col],
            "__date__": tmp["__date__"],
            "company": tmp[company_c] if company_c else None,
            "title": tmp[title_c] if title_c else None,
            "location": tmp[location_c] if location_c else None,
            "link": tmp[link_c] if link_c else None,
            "source_file": str(p)
        }
        df_sel = pd.DataFrame(sel)
        apps_frames.append(df_sel)
        picked_files.append(str(p))

    # Fallback: curated job sample in data dir if nothing was found in export
    if not apps_frames and curated_dir:
        for name in ["job_applications_2025_sample.csv", "job_applications_2025_monthly_counts.csv"]:
            p = curated_dir / name
            if p.exists():
                try:
                    df = robust_read_csv(p)
                    date_col = detect_date_column(df)
                    if date_col:
                        df["__date__"] = pd.to_datetime(df[date_col], errors="coerce")
                        df = df.dropna(subset=["__date__"]).sort_values("__date__")
                        if not df.empty:
                            apps_frames.append(df[[date_col]].rename(columns={date_col: "applied_date"}).assign(__date__=df["__date__"]))
                            picked_files.append(str(p))
                except Exception:
                    continue

    if not apps_frames:
        summary = {"status": "skipped", "reason": "no_applications_table_found"}
        (op / "output" / "applications_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
        return summary

    apps = pd.concat(apps_frames, ignore_index=True)
    if apps.empty:
        summary = {"status": "skipped", "reason": "no_valid_dates"}
        (op / "output" / "applications_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
        return summary

    # Metrics
    apps["year"] = apps["__date__"].dt.year
    by_year = apps.groupby("year").size().rename("applications")

    # Application periods (7-day gap)
    dates = apps["__date__"].tolist()
    dates_sorted = sorted(dates)
    periods: List[List[datetime]] = []
    if dates_sorted:
        cur = [dates_sorted[0]]
        for d in dates_sorted[1:]:
            if (d - cur[-1]).days > 7:
                periods.append(cur)
                cur = [d]
            else:
                cur.append(d)
        periods.append(cur)

    period_rows = []
    for i, grp in enumerate(periods, start=1):
        start = grp[0]
        end = grp[-1]
        num_days = max(1, (end - start).days + 1)
        num_apps = len(grp)
        period_rows.append({
            "period_index": i,
            "start_date": start.date().isoformat(),
            "end_date": end.date().isoformat(),
            "num_days": num_days,
            "num_applications": num_apps,
            "applications_per_day": round(num_apps / num_days, 4),
        })

    # Persist
    out_dir = op / "output"
    by_year.to_csv(out_dir / "applications_by_year.csv", header=True)
    pd.DataFrame(period_rows).to_csv(out_dir / "application_periods.csv", index=False)
    # Save raw applications events for deliverables and per-year sheets
    try:
        apps_out = op / "output" / "enriched" / "applications_events.xlsx"
        with pd.ExcelWriter(apps_out, engine="openpyxl") as w:
            apps[["applied_date","__date__","company","title","location","link","source_file"]].to_excel(w, index=False, sheet_name="Sheet1")
        # Per-year exports (focus on 2022 and 2025)
        for yr in (2022, 2025):
            sub = apps[apps["year"] == yr]
            if not sub.empty:
                out_y = op / "output" / "enriched" / f"Applications_{yr}.xlsx"
                with pd.ExcelWriter(out_y, engine="openpyxl") as wy:
                    sub[["applied_date","company","title","location","link","source_file"]].to_excel(wy, index=False, sheet_name="Applications")
    except Exception:
        pass

    summary = {
        "status": "success",
        "input_files": picked_files,
        "applications_by_year": by_year.to_dict(),
        "periods": period_rows,
    }
    (out_dir / "applications_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    return summary


def export_unified_workbook(op: Path, curated_path: Path, merged_path: Path, apps_summary: Dict[str, object]) -> Path:
    out_path = op / "output" / "enriched" / "unified_report.xlsx"
    # Load sheets
    curated_df = pd.read_excel(curated_path) if curated_path.exists() else pd.DataFrame()
    merged_df = pd.read_excel(merged_path) if merged_path.exists() else pd.DataFrame()

    # Summaries
    interviews_by_year = pd.Series(dtype=int)
    interviews_by_year_month = pd.DataFrame()
    by_medium = pd.DataFrame()
    all_events_by_year = pd.Series(dtype=int)
    events_by_type_year = pd.DataFrame()
    if not merged_df.empty and "event_date" in merged_df.columns:
        merged_df["event_date"] = pd.to_datetime(merged_df["event_date"], errors="coerce")
        if "event_year" not in merged_df.columns:
            merged_df["event_year"] = merged_df["event_date"].dt.year
        merged_non_null_year = merged_df.dropna(subset=["event_year"]).copy()
        merged_non_null_year["event_year"] = merged_non_null_year["event_year"].astype(int)
        # Interviews = curated events (event_type == interview). Outreach kept separate
        interviews_df = merged_non_null_year[merged_non_null_year.get("event_type").fillna("") == "interview"].copy()
        interviews_by_year = interviews_df.groupby("event_year").size().rename("interviews").astype(int)
        # Monthly uses only rows with valid dates
        monthly_df = merged_df.dropna(subset=["event_date"]).copy()
        monthly_df = monthly_df[monthly_df.get("event_type").fillna("") == "interview"]
        monthly_df["year"] = monthly_df["event_date"].dt.year
        interviews_by_year_month = (
            monthly_df.assign(month=monthly_df["event_date"].dt.to_period("M").astype(str))
            .groupby(["year", "month"]).size().rename("interviews").reset_index()
        )
        if "medium" in merged_df.columns:
            by_medium = interviews_df.groupby(["event_year", "medium"]).size().rename("interviews").reset_index().rename(columns={"event_year": "year"})
        # Outreach-inclusive totals
        all_events_by_year = merged_non_null_year.groupby("event_year").size().rename("events").astype(int)
        if "event_type" in merged_non_null_year.columns:
            events_by_type_year = (
                merged_non_null_year.groupby(["event_year","event_type"]).size().rename("events").reset_index()
                .rename(columns={"event_year": "year"})
            )

    # Applications frame from summary
    apps_by_year_df = pd.DataFrame()
    periods_df = pd.DataFrame()
    if apps_summary and apps_summary.get("status") == "success":
        aby = apps_summary.get("applications_by_year") or {}
        apps_by_year_df = pd.DataFrame({"year": list(aby.keys()), "applications": list(aby.values())})
        periods_df = pd.DataFrame(apps_summary.get("periods") or [])
        # Normalize year dtype
        try:
            apps_by_year_df["year"] = apps_by_year_df["year"].astype(int)
            apps_by_year_df["applications"] = apps_by_year_df["applications"].astype(int)
        except Exception:
            pass

    # Conversion metrics
    metrics_rows: List[Dict[str, object]] = []
    combined_years = sorted(set(interviews_by_year.index.tolist()) | set(apps_by_year_df.get("year", pd.Series([], dtype=int)).tolist()))
    for y in combined_years:
        interviews = int(interviews_by_year.get(y, 0)) if not interviews_by_year.empty else 0
        apps = int(apps_by_year_df[apps_by_year_df["year"] == y]["applications"].sum()) if not apps_by_year_df.empty else 0
        conv = (interviews / apps) if apps else None
        metrics_rows.append({"year": y, "applications": apps, "interviews": interviews, "interviews_per_application": conv})
    if metrics_rows:
        metrics_df = pd.DataFrame(metrics_rows).sort_values("year")
    else:
        metrics_df = pd.DataFrame(columns=["year", "applications", "interviews", "interviews_per_application"]) 

    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        if not curated_df.empty:
            curated_df.to_excel(w, index=False, sheet_name="Curated_Normalized")
        if not merged_df.empty:
            merged_df.to_excel(w, index=False, sheet_name="Merged_Events")
        if not interviews_by_year.empty:
            interviews_by_year.to_frame().reset_index().to_excel(w, index=False, sheet_name="Interviews_By_Year")
        if not interviews_by_year_month.empty:
            interviews_by_year_month.to_excel(w, index=False, sheet_name="Interviews_By_Month")
        if not by_medium.empty:
            by_medium.to_excel(w, index=False, sheet_name="Interviews_By_Medium")
        if not apps_by_year_df.empty:
            apps_by_year_df.to_excel(w, index=False, sheet_name="Applications_By_Year")
        if not periods_df.empty:
            periods_df.to_excel(w, index=False, sheet_name="Application_Periods")
        if not metrics_df.empty:
            metrics_df.to_excel(w, index=False, sheet_name="Summary_Metrics")

        # Outreach-inclusive summaries
        if not all_events_by_year.empty:
            all_events_by_year.to_frame().reset_index().rename(columns={"index":"year"}).to_excel(w, index=False, sheet_name="All_Events_By_Year")
        if not events_by_type_year.empty:
            events_by_type_year.to_excel(w, index=False, sheet_name="Events_By_Type_By_Year")

        # Queryable flat sheet combining all events (interview + outreach)
        all_cols = [
            "event_date","start_time","end_time","duration_minutes","company","role","medium","source","contacts","notes","period"
        ]
        flat = merged_df.copy()
        for c in ["event_year","event_type","__source__"]:
            if c not in flat.columns:
                flat[c] = None
        # Ensure columns exist
        for c in all_cols:
            if c not in flat.columns:
                flat[c] = None
        # Derive event_datetime for easier querying
        flat["event_datetime"] = pd.to_datetime(flat["event_date"], errors="coerce")
        # Backfill A: event_datetime and B: event_year from J: period when missing
        fallback_year = pd.to_numeric(flat.get("period"), errors="coerce")
        # event_year fallback
        flat["event_year"] = pd.to_numeric(flat.get("event_year"), errors="coerce")
        mask_year = flat["event_year"].isna() & fallback_year.notna()
        flat.loc[mask_year, "event_year"] = fallback_year[mask_year].astype(int)
        # event_datetime fallback uses Jan 1 of period when date is missing
        mask_dt = flat["event_datetime"].isna() & fallback_year.notna()
        if mask_dt.any():
            flat.loc[mask_dt, "event_datetime"] = pd.to_datetime(
                fallback_year[mask_dt].astype(int).astype(str) + "-01-01",
                errors="coerce"
            )
        # Reorder for readability
        flat_out = flat[[
            "event_datetime","event_year","event_type","company","contacts","role","medium","source","notes","period","start_time","end_time","duration_minutes","__source__"
        ]].copy()
        flat_out.sort_values(["event_year","event_datetime","company"], inplace=True)
        flat_out.to_excel(w, index=False, sheet_name="All_Events")

        # Build clean deliverable workbook
        out_path_clean = op / "output" / "enriched" / "unified_deliverable.xlsx"
        with pd.ExcelWriter(out_path_clean, engine="openpyxl") as wc:
            # Events
            flat_out.to_excel(wc, index=False, sheet_name="Events")
            # Applications raw if available
            apps_raw_path = op / "output" / "enriched" / "applications_events.xlsx"
            if apps_raw_path.exists():
                pd.read_excel(apps_raw_path).to_excel(wc, index=False, sheet_name="Applications_Raw")
            # Applications by year (2022, 2025) if present
            for yr in (2022, 2025):
                p_y = op / "output" / "enriched" / f"Applications_{yr}.xlsx"
                if p_y.exists():
                    pd.read_excel(p_y).to_excel(wc, index=False, sheet_name=f"Applications_{yr}")
            # Aggregates
            if not all_events_by_year.empty:
                all_events_by_year.to_frame().reset_index().rename(columns={"index":"year"}).to_excel(wc, index=False, sheet_name="All_Events_By_Year")
            if not events_by_type_year.empty:
                events_by_type_year.to_excel(wc, index=False, sheet_name="Events_By_Type_By_Year")
            # Inbound requests (outreach inbound) per year
            if not merged_df.empty:
                inbound = merged_df.copy()
                if "direction" in inbound.columns:
                    inbound = inbound[inbound["direction"].fillna("").str.lower() == "inbound"]
                inbound = inbound[inbound.get("event_type").fillna("") == "outreach"]
                inbound["event_year"] = inbound["event_year"].fillna(pd.to_datetime(inbound["event_date"], errors="coerce").dt.year)
                for yr in (2022, 2025):
                    sub = inbound[inbound["event_year"] == yr].copy()
                    if not sub.empty:
                        cols = [
                            "event_date","company","contacts","medium","source","notes","direction"
                        ]
                        for c in cols:
                            if c not in sub.columns:
                                sub[c] = None
                        sub = sub[cols]
                        sub.to_excel(wc, index=False, sheet_name=f"Inbound_Requests_{yr}")
            if not apps_by_year_df.empty:
                apps_by_year_df.to_excel(wc, index=False, sheet_name="Applications_By_Year")
            if not periods_df.empty:
                periods_df.to_excel(wc, index=False, sheet_name="Application_Periods")
            # QA errors
            try:
                errs = []
                # Missing event_datetime
                missing_dt = flat_out[flat_out["event_datetime"].isna()]
                for _, r in missing_dt.iterrows():
                    errs.append({"issue":"missing_event_datetime","company":r.get("company"),"contacts":r.get("contacts"),"medium":r.get("medium"),"source":r.get("source"),"notes":r.get("notes")})
                # Missing company & contacts
                missing_cc = flat_out[flat_out["company"].isna() & flat_out["contacts"].isna()]
                for _, r in missing_cc.iterrows():
                    errs.append({"issue":"missing_company_and_contacts","event_year":r.get("event_year"),"medium":r.get("medium"),"source":r.get("source")})
                # Applications QA: missing applied_date or company/title
                apps_raw = pd.read_excel(apps_raw_path) if apps_raw_path.exists() else pd.DataFrame()
                if not apps_raw.empty:
                    miss_app = apps_raw[apps_raw["applied_date"].isna() | (apps_raw.get("company").isna() & apps_raw.get("title").isna())]
                    for _, r in miss_app.iterrows():
                        errs.append({"issue":"applications_missing_core_fields","applied_date":r.get("applied_date"),"company":r.get("company"),"title":r.get("title"),"source_file":r.get("source_file")})
                if errs:
                    pd.DataFrame(errs).to_excel(wc, index=False, sheet_name="QA_Errors")
            except Exception:
                pass
            # Data Dictionary
            dd = pd.DataFrame([
                {"column":"event_datetime","description":"Parsed datetime of the event when available"},
                {"column":"event_year","description":"Event year from date or fallback when needed"},
                {"column":"event_type","description":"interview (curated meeting) or outreach (LinkedIn DM/Invitation)"},
                {"column":"company","description":"Counterparty or organization"},
                {"column":"contacts","description":"Primary contact involved"},
                {"column":"role","description":"Role/Purpose of the interaction"},
                {"column":"medium","description":"Channel (Google Meet, Teams, LinkedIn DM, etc.)"},
                {"column":"source","description":"Source system (Calendar, Gmail, LinkedIn, etc.)"},
                {"column":"notes","description":"Short notes"},
                {"column":"period","description":"Canonical period/year from source (e.g., 2022, 2025)"},
                {"column":"start_time","description":"Start time when available"},
                {"column":"end_time","description":"End time when available"},
                {"column":"duration_minutes","description":"Duration in minutes when provided"},
                {"column":"__source__","description":"Merged source tag (curated/linkedin)"},
            ])
            dd.to_excel(wc, index=False, sheet_name="Data_Dictionary")

        # Convenience: filtered merged list for 2025
        if not merged_df.empty and "event_year" in merged_df.columns:
            try:
                merged_2025 = merged_df.copy()
                if "event_year" not in merged_2025.columns:
                    merged_2025["event_year"] = pd.to_datetime(merged_2025["event_date"], errors="coerce").dt.year
                merged_2025 = merged_2025[merged_2025["event_year"] == 2025]
                if not merged_2025.empty:
                    merged_2025.to_excel(w, index=False, sheet_name="Merged_Events_2025")
            except Exception:
                pass

        # Reconciliation sheet to surface discrepancies between event_date-only vs fallback-year counts
        try:
            if not merged_df.empty:
                temp = merged_df.copy()
                temp["event_date"] = pd.to_datetime(temp["event_date"], errors="coerce")
                if "event_year" not in temp.columns:
                    temp["event_year"] = temp["event_date"].dt.year
                year_by_date = temp.dropna(subset=["event_date"]).assign(_y=temp["event_date"].dt.year).groupby("_y").size().rename("by_event_date").to_frame()
                year_by_fallback = temp.dropna(subset=["event_year"]).assign(_y=temp["event_year"].astype(int)).groupby("_y").size().rename("by_event_year").to_frame()
                recon = year_by_date.join(year_by_fallback, how="outer").fillna(0).astype(int).reset_index().rename(columns={"_y": "year"})
                recon.to_excel(w, index=False, sheet_name="Reconciliation")
        except Exception:
            pass

        # Year-specific overviews for 2022 and 2025
        years_to_render: List[int] = []
        if not interviews_by_year.empty or not apps_by_year_df.empty:
            for y in (2022, 2025):
                present_i = (not interviews_by_year.empty) and (y in set(interviews_by_year.index.tolist()))
                present_a = (not apps_by_year_df.empty) and (y in set(apps_by_year_df.get("year", pd.Series([], dtype=int)).tolist()))
                if present_i or present_a:
                    years_to_render.append(y)

        for y in years_to_render:
            sheet_name = f"Overview_{y}"
            start = 0
            # Interviews monthly for year
            interviews_monthly_y = pd.DataFrame()
            if not merged_df.empty and "event_date" in merged_df.columns:
                md_y = merged_df.copy()
                md_y["event_date"] = pd.to_datetime(md_y["event_date"], errors="coerce")
                md_y = md_y.dropna(subset=["event_date"]) 
                md_y = md_y[md_y["event_date"].dt.year == y]
                if not md_y.empty:
                    interviews_monthly_y = (
                        md_y.set_index("event_date").assign(count=1)["count"].resample("MS").sum().rename("interviews").to_frame()
                    )
                    interviews_monthly_y.index = interviews_monthly_y.index.strftime("%Y-%m")

            # Interviews by medium
            by_medium_y = pd.DataFrame()
            if not merged_df.empty and "medium" in merged_df.columns and "event_date" in merged_df.columns:
                mdm = merged_df.copy()
                mdm["event_date"] = pd.to_datetime(mdm["event_date"], errors="coerce")
                mdm = mdm.dropna(subset=["event_date"]) 
                mdm = mdm[mdm["event_date"].dt.year == y]
                if not mdm.empty:
                    by_medium_y = mdm.groupby("medium").size().rename("interviews").sort_values(ascending=False).reset_index()

            # Applications yearly and periods filtered for this year
            apps_total_y = pd.DataFrame()
            periods_y = pd.DataFrame()
            conv_row = pd.DataFrame()
            if apps_summary and apps_summary.get("status") == "success":
                aby = apps_summary.get("applications_by_year") or {}
                apps_y_val = int(aby.get(str(y), 0))
                apps_total_y = pd.DataFrame({"year": [y], "applications": [apps_y_val]})
                # Filter periods by year overlap
                periods = apps_summary.get("periods") or []
                if periods:
                    tmp = pd.DataFrame(periods)
                    tmp["start_date"] = pd.to_datetime(tmp["start_date"], errors="coerce")
                    tmp["end_date"] = pd.to_datetime(tmp["end_date"], errors="coerce")
                    sel = tmp[(tmp["start_date"].dt.year == y) | (tmp["end_date"].dt.year == y)]
                    periods_y = sel.copy()
                # Conversion row
                interviews_y_val = int(interviews_by_year.get(y, 0)) if not interviews_by_year.empty else 0
                conv = (interviews_y_val / apps_y_val) if apps_y_val else None
                conv_row = pd.DataFrame({"year": [y], "applications": [apps_y_val], "interviews": [interviews_y_val], "interviews_per_application": [conv]})

            # Write blocks into one sheet using offsets
            if not interviews_monthly_y.empty:
                interviews_monthly_y.reset_index().rename(columns={"index": "month"}).to_excel(w, index=False, sheet_name=sheet_name, startrow=start)
                start += len(interviews_monthly_y) + 3
            if not by_medium_y.empty:
                by_medium_y.to_excel(w, index=False, sheet_name=sheet_name, startrow=start)
                start += len(by_medium_y) + 3
            if not apps_total_y.empty:
                apps_total_y.to_excel(w, index=False, sheet_name=sheet_name, startrow=start)
                start += len(apps_total_y) + 3
            if not periods_y.empty:
                periods_y.to_excel(w, index=False, sheet_name=sheet_name, startrow=start)
                start += len(periods_y) + 3
            if not conv_row.empty:
                conv_row.to_excel(w, index=False, sheet_name=sheet_name, startrow=start)
    try:
        wb = load_workbook(str(out_path))
        if "PROVENANCE" in wb.sheetnames:
            wb.remove(wb["PROVENANCE"])
        # Append PROVENANCE as last sheet
        ws = wb.create_sheet("PROVENANCE")
        ws["A1"] = "operation_id"; ws["B1"] = op.name
        ws["A2"] = "generated_at_utc"; ws["B2"] = datetime.utcnow().isoformat() + "Z"
        ws["A3"] = "output_path"; ws["B3"] = str(out_path)
        # Basic number/date formats for key sheets
        def fmt_date(ws, col_letter: str):
            try:
                for r in range(2, ws.max_row + 1):
                    cell = ws[f"{col_letter}{r}"]
                    if isinstance(cell.value, datetime):
                        cell.number_format = "yyyy-mm-dd"
            except Exception:
                pass

        for name in ["Curated_Normalized", "Merged_Events"]:
            if name in wb.sheetnames:
                sh = wb[name]
                # Detect "event_date" column index
                headers = {sh[f"{chr(65+i)}1"].value: chr(65+i) for i in range(sh.max_column)}
                col = headers.get("event_date") or headers.get("applied_date")
                if col:
                    fmt_date(sh, col)

        if "Application_Periods" in wb.sheetnames:
            sh = wb["Application_Periods"]
            headers = {sh[f"{chr(65+i)}1"].value: chr(65+i) for i in range(sh.max_column)}
            for key in ("start_date", "end_date"):
                col = headers.get(key)
                if col:
                    fmt_date(sh, col)
        wb.save(str(out_path))
    except Exception:
        pass
    return out_path


# -----------------------------
# CLI
# -----------------------------


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Interviews pipeline: normalize, enrich, merge and summarize (2022 vs 2025)")
    p.add_argument("--data-dir", default="Data Tests/Interviews", help="Directory with curated CSVs and LinkedIn export zips")
    p.add_argument("--artifacts-root", default="artifacts/operations", help="Artifacts root directory")
    p.add_argument("--mode", choices=["minimal","full"], default="minimal", help="Output mode: minimal (2 datasets) or full reports")
    return p.parse_args()


def main() -> None:
    args = parse_args()
    curated_dir = Path(args.data_dir)
    artifacts_root = Path(args.artifacts_root)

    op = ensure_op_dir(artifacts_root)
    write_audit(op, agent="interviews-pipeline", inputs={"curated_dir": str(curated_dir)}, note="initialized")

    # Step 1: Extract and inventory LinkedIn export (if available)
    export_root = extract_linkedin_export(curated_dir, op)
    catalog = inventory_export(export_root, op / "processing", op / "output", minimal=(args.mode == "minimal")) if export_root else {}

    outputs: Dict[str, Optional[str]] = {"operation_id": op.name}
    if args.mode == "minimal":
        # In-memory normalization
        curated_df = normalize_curated_df(curated_dir, export_root)
        li_df = build_linkedin_events_df(export_root)
        merged_df = merge_events_df(curated_df, li_df)
        apps_df, apps_summary = extract_applications_events_df(export_root, curated_dir)
        two = export_minimal_deliverables(op, merged_df, apps_df)
        outputs.update({
            "interviews_dataset": two.get("interviews_dataset"),
            "applications_dataset": two.get("applications_dataset"),
        })
    else:
        # Full outputs path (backward compatible)
        curated_norm_path = normalize_curated(curated_dir, export_root, op)
        linkedin_events_path = build_linkedin_events(export_root, op)
        merged_path = merge_events(op, curated_norm_path, linkedin_events_path)
        apps_summary = detect_and_analyze_applications(export_root, op, curated_dir)
        unified_path = export_unified_workbook(op, curated_norm_path, merged_path, apps_summary)
        outputs.update({
            "normalized_interviews": str(curated_norm_path),
            "linkedin_events": str(linkedin_events_path) if linkedin_events_path else None,
            "merged_events": str(merged_path),
            "unified_report": str(unified_path),
            "applications_summary": str(op / "output" / "applications_summary.json"),
            "dataset_catalog": str(op / "output" / "dataset_catalog.xlsx"),
        })

    # Finalize audit
    write_audit(
        op,
        agent="interviews-pipeline",
        inputs={
            "curated_dir": str(curated_dir),
            "export_root": str(export_root) if export_root else None,
        },
        note="completed",
    )

    # Print paths for quick reference
    print(json.dumps(outputs, indent=2))


if __name__ == "__main__":
    main()


