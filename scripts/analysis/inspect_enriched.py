import json
import sys
from pathlib import Path
from openpyxl import load_workbook


def inspect_xlsx(path: Path):
    try:
        wb = load_workbook(str(path), read_only=True)
        sheets = wb.sheetnames
        out = {"sheets": sheets, "first_sheet": sheets[0] if sheets else None}
        if sheets:
            ws = wb[sheets[0]]
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            first_rows = []
            for r in range(2, min(7, ws.max_row + 1)):
                row = [ws.cell(r, c).value for c in range(1, min(11, ws.max_column + 1))]
                first_rows.append(row)
            out.update({"headers": headers, "sample": first_rows})
        return out
    except Exception as e:
        return {"error": str(e)}


def main() -> None:
    ops_root = Path('artifacts/operations')
    candidates = [p for p in ops_root.iterdir() if p.is_dir() and p.name[:8].isdigit()]
    latest = max(candidates, key=lambda p: p.name) if candidates else max([p for p in ops_root.iterdir() if p.is_dir()], key=lambda p: p.name)
    root = latest / 'output' / 'enriched'
    files = [
        root / 'merged_interview_events.xlsx',
        root / 'unified_report.xlsx',
        root / 'normalized_interviews.xlsx',
        root / 'linkedin_events.xlsx',
    ]
    res = {}
    for f in files:
        res[f.name] = inspect_xlsx(f)
    print(json.dumps(res, indent=2, default=str))


if __name__ == '__main__':
    main()


