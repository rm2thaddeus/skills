import argparse
import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import unicodedata
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


@dataclass
class OperationPaths:
    op_dir: Path
    input_dir: Path
    processing_dir: Path
    output_dir: Path


def ensure_operation_dirs(output_root: Path) -> OperationPaths:
    op_id = datetime.utcnow().strftime("%Y%m%d-%H%M%S-%f")
    op_dir = output_root / op_id
    input_dir = op_dir / "input"
    processing_dir = op_dir / "processing"
    output_dir = op_dir / "output"
    input_dir.mkdir(parents=True, exist_ok=True)
    processing_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    return OperationPaths(op_dir, input_dir, processing_dir, output_dir)


def detect_header_row(df_no_header: pd.DataFrame, look_rows: int = 10) -> Optional[int]:
    for i in range(min(look_rows, len(df_no_header))):
        row_values = df_no_header.iloc[i].astype(str).str.strip().tolist()
        if any(v.lower() == "identifiant" for v in row_values):
            return i
    return None


def normalize_header(value: Any) -> str:
    s = str(value or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def strip_accents(text: str) -> str:
    try:
        text = str(text)
    except Exception:
        return ""
    return "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))


def clean_medecins(df: pd.DataFrame) -> pd.DataFrame:
    # Drop columns that are entirely NaN
    df = df.dropna(axis=1, how="all")
    # Normalize headers
    df.columns = [normalize_header(c) for c in df.columns]
    # Drop fully empty rows
    df = df.dropna(how="all")
    # Strip strings
    for col in df.select_dtypes(include=["object"]).columns:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace({"nan": None})
    return df


def parse_duration_minutes(text: Any) -> Optional[float]:
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return None
    s = str(text).strip().lower()
    m = re.match(r"^(\d+(?:[\.,]\d+)?)\s*min", s)
    if m:
        return float(m.group(1).replace(",", "."))
    try:
        return float(s)
    except Exception:
        return None


def excel_serial_to_datetime(val: Any) -> Optional[pd.Timestamp]:
    try:
        return pd.to_datetime(val, origin="1899-12-30", unit="D", errors="coerce")
    except Exception:
        return None


def clean_visites(df: pd.DataFrame) -> pd.DataFrame:
    df = df.dropna(axis=1, how="all")
    df.columns = [normalize_header(c) for c in df.columns]
    df = df.dropna(how="all").copy()
    # Strip strings
    for col in df.select_dtypes(include=["object"]).columns:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace({"nan": None})
    # Date conversion
    date_col_candidates = [c for c in df.columns if c.lower().startswith("date")]
    if date_col_candidates:
        dc = date_col_candidates[0]
        df[dc] = pd.to_numeric(df[dc], errors="coerce")
        df[dc] = pd.to_datetime(df[dc], origin="1899-12-30", unit="D", errors="coerce")
        df.rename(columns={dc: "Date visite"}, inplace=True)
    # Duration minutes
    dur_candidates = [c for c in df.columns if "dur" in c.lower()]
    if dur_candidates:
        dcol = dur_candidates[0]
        df["Duree_visite_min"] = df[dcol].apply(parse_duration_minutes)
    return df


def read_instructions_to_markdown(xlsx_path: Path, output_md: Path) -> None:
    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    if "Instructions" not in wb.sheetnames:
        return
    ws = wb["Instructions"]
    lines: List[str] = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 200, 200), values_only=True):
        vals = [str(v).strip() for v in row if v is not None and str(v).strip() != ""]
        if vals:
            lines.append(" ".join(vals))
    output_md.write_text("\n".join(lines), encoding="utf-8")


def map_headers_to_columns(xlsx_path: Path, sheet_name: str, header_row_index_1based: int) -> Dict[str, str]:
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    ws = wb[sheet_name]
    mapping: Dict[str, str] = {}
    for cell in ws[header_row_index_1based]:
        val = cell.value
        if val is None:
            continue
        header = normalize_header(val)
        mapping[header] = get_column_letter(cell.column)
    return mapping


def propose_dashboard_formulas(
    xlsx_path: Path,
    visites_sheet: str,
    header_row_visites_1based: int,
    header_to_col_visites: Dict[str, str],
    medecins_sheet: str,
    header_row_medecins_1based: int,
    header_to_col_medecins: Dict[str, str],
    modes: List[str],
    secteurs: List[str],
    ars_values: List[str],
    region_values: List[str],
    specialite_values: List[str],
    delegues: List[str],
) -> Dict[str, Any]:
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    if "Dashboard" not in wb.sheetnames:
        return {"status": "no_dashboard_sheet"}
    ws = wb["Dashboard"]

    def data_range_vis(col_letter: str) -> str:
        start = header_row_visites_1based + 1
        return f"'{visites_sheet}'!${col_letter}${start}:${col_letter}$1048576"

    def data_range_med(col_letter: str) -> str:
        start = header_row_medecins_1based + 1
        return f"'{medecins_sheet}'!${col_letter}${start}:${col_letter}$1048576"

    proposals: Dict[str, Any] = {"cells": []}
    # Build quick lookup of labels on Dashboard
    cell_values: Dict[str, Tuple[int, int]] = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(ws.max_column, 60) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            try:
                text = str(v).strip()
            except Exception:
                continue
            if not text:
                continue
            key = normalize_header(text).lower()
            cell_values[key] = (r, c)

    def find_and_assign(label_keywords: List[str], formula_builder) -> None:
        for key, (r, c) in list(cell_values.items()):
            key_ascii = strip_accents(key)
            if all((k in key) or (k in key_ascii) for k in label_keywords):
                target_r, target_c = r, c + 1  # write to the right cell
                excel_a1 = f"{get_column_letter(target_c)}{target_r}"
                formula = formula_builder((r, c), excel_a1)
                if formula:
                    proposals["cells"].append({
                        "label_cell": f"{get_column_letter(c)}{r}",
                        "target_cell": excel_a1,
                        "label_text": key,
                        "formula": formula,
                    })
                break

    ident_vis_col = header_to_col_visites.get("Identifiant")
    ident_med_col = header_to_col_medecins.get("Identifiant")
    mode_col = None
    for k in header_to_col_visites.keys():
        if "mode" in k.lower():
            mode_col = header_to_col_visites[k]
            break
    secteur_col = header_to_col_visites.get("Secteur")

    # Total visits
    if ident_vis_col:
        find_and_assign(["total", "visite"], lambda _label, _a1: f"=COUNTA({data_range_vis(ident_vis_col)})")
    # Unique doctors visited
    if ident_vis_col:
        find_and_assign(["medecin", "visite"], lambda _label, _a1: f"=COUNTA(UNIQUE({data_range_vis(ident_vis_col)}))")

    # Mode counts
    if mode_col:
        for _key, (r, c) in list(cell_values.items()):
            original_text = ws.cell(row=r, column=c).value
            if not isinstance(original_text, str):
                continue
            val = original_text.strip()
            if val in modes:
                target_c = c + 1
                excel_a1 = f"{get_column_letter(target_c)}{r}"
                formula = f"=COUNTIF({data_range_vis(mode_col)},{get_column_letter(c)}{r})"
                proposals["cells"].append({
                    "label_cell": f"{get_column_letter(c)}{r}",
                    "target_cell": excel_a1,
                    "label_text": val,
                    "formula": formula,
                })

    # Secteur counts (from visites)
    if secteur_col:
        for _key, (r, c) in list(cell_values.items()):
            original_text = ws.cell(row=r, column=c).value
            if not isinstance(original_text, str):
                continue
            val = original_text.strip()
            if val in secteurs:
                target_c = c + 1
                excel_a1 = f"{get_column_letter(target_c)}{r}"
                formula = f"=COUNTIF({data_range_vis(secteur_col)},{get_column_letter(c)}{r})"
                proposals["cells"].append({
                    "label_cell": f"{get_column_letter(c)}{r}",
                    "target_cell": excel_a1,
                    "label_text": val,
                    "formula": formula,
                })

    # ARS / Region / Specialite counts (via medecins lookup)
    def add_lookup_count_for_values(values: List[str], med_attr_key: str) -> None:
        if not (ident_vis_col and ident_med_col):
            return
        med_attr_col = header_to_col_medecins.get(med_attr_key)
        if not med_attr_col:
            return
        for _key, (r, c) in list(cell_values.items()):
            original_text = ws.cell(row=r, column=c).value
            if not isinstance(original_text, str):
                continue
            val = original_text.strip()
            if val in values:
                target_c = c + 1
                excel_a1 = f"{get_column_letter(target_c)}{r}"
                formula = (
                    f"=SUMPRODUCT(--(XLOOKUP({data_range_vis(ident_vis_col)},{data_range_med(ident_med_col)}"
                    f",{data_range_med(med_attr_col)})={get_column_letter(c)}{r}))"
                )
                proposals["cells"].append({
                    "label_cell": f"{get_column_letter(c)}{r}",
                    "target_cell": excel_a1,
                    "label_text": val,
                    "formula": formula,
                })

    if ars_values:
        add_lookup_count_for_values(ars_values, "ARS")
    if region_values:
        add_lookup_count_for_values(region_values, "Région")
    # Spécialités: check either Spécialité 1 or Spécialité 2; use OR by SUM of two comparisons
    if specialite_values and ident_vis_col and ident_med_col:
        spec1_col = header_to_col_medecins.get("Spécialité 1")
        spec2_col = header_to_col_medecins.get("Spécialité 2")
        for _key, (r, c) in list(cell_values.items()):
            original_text = ws.cell(row=r, column=c).value
            if not isinstance(original_text, str):
                continue
            val = original_text.strip()
            if val in specialite_values:
                target_c = c + 1
                excel_a1 = f"{get_column_letter(target_c)}{r}"
                parts = []
                if spec1_col:
                    parts.append(
                        f"--(XLOOKUP({data_range_vis(ident_vis_col)},{data_range_med(ident_med_col)},{data_range_med(spec1_col)})={get_column_letter(c)}{r})"
                    )
                if spec2_col:
                    parts.append(
                        f"--(XLOOKUP({data_range_vis(ident_vis_col)},{data_range_med(ident_med_col)},{data_range_med(spec2_col)})={get_column_letter(c)}{r})"
                    )
                if parts:
                    joined = "+".join(parts)
                    formula = f"=SUMPRODUCT({joined})"
                    proposals["cells"].append({
                        "label_cell": f"{get_column_letter(c)}{r}",
                        "target_cell": excel_a1,
                        "label_text": val,
                        "formula": formula,
                    })

    # Délégué counts (from visites)
    delegue_col_letter = None
    for k, v in header_to_col_visites.items():
        if "delegue" in strip_accents(k).lower():
            delegue_col_letter = v
            break
    if delegue_col_letter:
        for _key, (r, c) in list(cell_values.items()):
            original_text = ws.cell(row=r, column=c).value
            if not isinstance(original_text, str):
                continue
            val = original_text.strip()
            if val in delegues:
                target_c = c + 1
                excel_a1 = f"{get_column_letter(target_c)}{r}"
                formula = f"=COUNTIF({data_range_vis(delegue_col_letter)},{get_column_letter(c)}{r})"
                proposals["cells"].append({
                    "label_cell": f"{get_column_letter(c)}{r}",
                    "target_cell": excel_a1,
                    "label_text": val,
                    "formula": formula,
                })

    proposals["status"] = "proposed"
    proposals["notes"] = "Formulas are proposed only; application requires approval."
    return proposals


def compute_kpis(visites: pd.DataFrame) -> Dict[str, Any]:
    kpis: Dict[str, Any] = {}
    kpis["total_visites"] = int(len(visites))
    if "Identifiant" in visites.columns:
        kpis["unique_medecins_visites"] = int(visites["Identifiant"].nunique())
    if "Mode de visite" in visites.columns:
        by_mode = visites["Mode de visite"].value_counts().rename_axis("mode").reset_index(name="visites")
        kpis["visites_par_mode"] = by_mode.to_dict(orient="records")
    if "Secteur" in visites.columns:
        by_sec = visites["Secteur"].value_counts().rename_axis("secteur").reset_index(name="visites")
        kpis["visites_par_secteur"] = by_sec.to_dict(orient="records")
    if "Date visite" in visites.columns and pd.api.types.is_datetime64_any_dtype(visites["Date visite"]):
        monthly = (
            visites.set_index("Date visite").assign(count=1)["count"].resample("MS").sum()
        )
        kpis["visites_par_mois"] = monthly.reset_index().assign(
            month=lambda d: d["Date visite"].dt.strftime("%Y-%m")
        )[["month", "count"]].to_dict(orient="records")
    if "Duree_visite_min" in visites.columns:
        kpis["duree_moyenne_min"] = float(pd.to_numeric(visites["Duree_visite_min"], errors="coerce").mean())
    delegue_col = next((c for c in visites.columns if "delegue" in strip_accents(c).lower()), None)
    if delegue_col:
        top_del = visites[delegue_col].value_counts().rename_axis("delegue").reset_index(name="visites")
        kpis["visites_par_delegue"] = top_del.to_dict(orient="records")
    return kpis


def run(input_path: Path, output_root: Path) -> Dict[str, Any]:
    paths = ensure_operation_dirs(output_root)
    # Reference input
    (paths.input_dir / "README.txt").write_text(f"Input referenced at: {input_path}\n", encoding="utf-8")

    # Read all sheets without headers to detect header rows
    med_no_header = pd.read_excel(input_path, sheet_name=1, header=None)
    vis_no_header = pd.read_excel(input_path, sheet_name=2, header=None)

    med_header_idx = detect_header_row(med_no_header) or 2
    vis_header_idx = detect_header_row(vis_no_header) or 2

    med = pd.read_excel(input_path, sheet_name=1, header=med_header_idx)
    vis = pd.read_excel(input_path, sheet_name=2, header=vis_header_idx)

    med = clean_medecins(med)
    vis = clean_visites(vis)

    # Persist cleaned data
    med.to_csv(paths.output_dir / "medecins_clean.csv", index=False, encoding="utf-8")
    vis.to_csv(paths.output_dir / "visites_clean.csv", index=False, encoding="utf-8")

    # KPIs
    kpis = compute_kpis(vis)
    (paths.output_dir / "kpis.json").write_text(json.dumps(kpis, indent=2, ensure_ascii=False), encoding="utf-8")

    # Report
    report_lines: List[str] = []
    report_lines.append("## Résultats du test de dashboarding\n")
    report_lines.append(f"- Total visites: {kpis.get('total_visites', 0)}")
    if "unique_medecins_visites" in kpis:
        report_lines.append(f"- Médecins visités (uniques): {kpis['unique_medecins_visites']}")
    if "duree_moyenne_min" in kpis and kpis["duree_moyenne_min"]:
        report_lines.append(f"- Durée moyenne (min): {kpis['duree_moyenne_min']:.1f}")
    report_md = paths.output_dir / "report.md"
    report_md.write_text("\n".join(report_lines) + "\n", encoding="utf-8")

    # Save distributions as CSVs
    if "Mode de visite" in vis.columns:
        vis["Mode de visite"].value_counts().rename_axis("mode").reset_index(name="visites").to_csv(
            paths.output_dir / "visites_par_mode.csv", index=False, encoding="utf-8"
        )
    if "Secteur" in vis.columns:
        vis["Secteur"].value_counts().rename_axis("secteur").reset_index(name="visites").to_csv(
            paths.output_dir / "visites_par_secteur.csv", index=False, encoding="utf-8"
        )

    # Instructions extraction
    read_instructions_to_markdown(input_path, paths.processing_dir / "instructions.md")

    # Header to column mapping for formula building
    header_to_col_visites = map_headers_to_columns(
        input_path,
        sheet_name="Données visites",
        header_row_index_1based=vis_header_idx + 1,
    )
    header_to_col_medecins = map_headers_to_columns(
        input_path,
        sheet_name="Données médecins",
        header_row_index_1based=med_header_idx + 1,
    )
    (paths.processing_dir / "column_map_visites.json").write_text(
        json.dumps(header_to_col_visites, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    (paths.processing_dir / "column_map_medecins.json").write_text(
        json.dumps(header_to_col_medecins, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    modes = []
    if "Mode de visite" in vis.columns:
        modes = [m for m in vis["Mode de visite"].dropna().unique().tolist() if isinstance(m, str)]
    secteurs = []
    if "Secteur" in vis.columns:
        secteurs = [s for s in vis["Secteur"].dropna().unique().tolist() if isinstance(s, str)]

    ars_values = []
    if "ARS" in med.columns:
        ars_values = [a for a in med["ARS"].dropna().unique().tolist() if isinstance(a, str)]
    region_values = []
    if "Région" in med.columns:
        region_values = [r for r in med["Région"].dropna().unique().tolist() if isinstance(r, str)]
    specialite_values = []
    for col in [c for c in med.columns if c.startswith("Spécialité")]:
        specialite_values.extend([s for s in med[col].dropna().unique().tolist() if isinstance(s, str)])
    specialite_values = sorted(set(specialite_values))

    delegues = []
    del_col = next((c for c in vis.columns if "delegue" in strip_accents(c).lower()), None)
    if del_col:
        delegues = [d for d in vis[del_col].dropna().unique().tolist() if isinstance(d, str)]

    proposals = propose_dashboard_formulas(
        xlsx_path=input_path,
        visites_sheet="Données visites",
        header_row_visites_1based=vis_header_idx + 1,
        header_to_col_visites=header_to_col_visites,
        medecins_sheet="Données médecins",
        header_row_medecins_1based=med_header_idx + 1,
        header_to_col_medecins=header_to_col_medecins,
        modes=modes,
        secteurs=secteurs,
        ars_values=ars_values,
        region_values=region_values,
        specialite_values=specialite_values,
        delegues=delegues,
    )
    (paths.processing_dir / "proposed_formulas.json").write_text(
        json.dumps(proposals, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    # Human-readable formulas
    formulas_md = paths.output_dir / "formulas.md"
    lines = ["## Formules proposées (non appliquées)\n"]
    for cell in proposals.get("cells", []):
        lines.append(f"- {cell['target_cell']}: {cell['formula']} (label: {cell['label_cell']})")
    formulas_md.write_text("\n".join(lines) + "\n", encoding="utf-8")

    # Audit
    audit = {
        "operation_id": paths.op_dir.name,
        "timestamp": datetime.utcnow().isoformat() + "Z",
        "agent": "excel-processing",
        "input_file": str(input_path),
        "output_dir": str(paths.output_dir),
        "processing_artifacts": [
            str(paths.processing_dir / "instructions.md"),
            str(paths.processing_dir / "column_map_visites.json"),
            str(paths.processing_dir / "column_map_medecins.json"),
            str(paths.processing_dir / "proposed_formulas.json"),
        ],
        "changes": [
            "Data cleaned (headers normalized, dates parsed, durations extracted)",
            "KPIs computed and saved",
            "Formulas proposed for Dashboard sheet (including ARS/Région/Spécialités/Délégué)",
        ],
        "validation": {
            "status": "success",
            "notes": "Read-only processing; no workbook modifications yet",
        },
        "approval_status": "pending",
    }
    (paths.op_dir / "audit.json").write_text(json.dumps(audit, indent=2), encoding="utf-8")

    summary = {
        "status": "success",
        "operation_id": paths.op_dir.name,
        "input_file": str(input_path),
        "artifacts": {
            "kpis_json": str(paths.output_dir / "kpis.json"),
            "report_md": str(paths.output_dir / "report.md"),
            "medecins_clean_csv": str(paths.output_dir / "medecins_clean.csv"),
            "visites_clean_csv": str(paths.output_dir / "visites_clean.csv"),
            "visites_par_mode_csv": str(paths.output_dir / "visites_par_mode.csv"),
            "visites_par_secteur_csv": str(paths.output_dir / "visites_par_secteur.csv"),
            "formulas_md": str(paths.output_dir / "formulas.md"),
            "proposed_formulas_json": str(paths.processing_dir / "proposed_formulas.json"),
        },
    }
    (paths.output_dir / "summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    return summary


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Process dashboarding test workbook and compute KPIs")
    p.add_argument("input_path", type=str, help="Path to the dashboarding test .xlsx")
    p.add_argument("--output-root", type=str, default="artifacts/operations", help="Artifacts root directory")
    return p.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input_path)
    output_root = Path(args.output_root)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    summary = run(input_path, output_root)
    print(json.dumps(summary, ensure_ascii=False))


if __name__ == "__main__":
    main()


